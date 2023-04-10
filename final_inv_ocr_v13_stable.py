
# -*- coding: utf-8 -*-
# 程序名： final_inv_ocr
# Author: AllenZhang.张海林

# 程序功能：电子发票转成图片并识别商品明细
# v9版：可以先旋转角度。先截二维码识别票号金额。以大写金额作为原点坐标截取其他区域小图识别，以openpyxl操作excel。
# 可选择用识别引擎：快速-mb 平衡:sv 精细-pp  (总体上，预识别用mb，精细用pd，速度和精确度比较好。)
# 增加重复发票识别，跳过已扫描的重复发票。
# 适配大批量发票识别，边识别边存储。适配断续工作，检查result文件时，检查发票路径是否已全部识别。
# 改为固定位置切图，适配固定的扫描仪。
# acrobat pro dc转换隐藏软件界面：1.首选项->一般->警告->不显示  2.首选项->安全性和安全性增强 全部取消
# win7 x64安装vscode 1.69.0版。再高的版本不兼容win7。官网下载路径中的网址替换为中文镜像网址后下载可以提速。

# log: 2022.11.14 create
#      2022.12.8  提速，多线程优化。
#      2022.12.9  修改显示日期问题。
#      2022.12.21 添加电子票通用识别。转为excel，先识别里面图片和二维码，然后匹配表格内容。              
import numba as nb
import imghdr
import math
import os
import re
import shutil
from collections import OrderedDict
from datetime import datetime
from glob import glob
from pathlib import Path
from tkinter import filedialog
from tkinter import Tk
import cv2
import numpy as np
import paddlehub as hub
import pandas as pd
import psutil
from openpyxl import cell, load_workbook
from openpyxl.styles import Font, colors
from paddleocr import PaddleOCR, draw_ocr
from PIL import Image, ImageDraw, ImageEnhance, ImageFont
from pyzbar import pyzbar
from zipfile import ZipFile
import fitz #pip install pymupdf

# --------------------------Func walk_folder_ocr()---------------------------
# 遍历文件夹内的发票文件，识别。
@nb.autojit
def walk_folder_ocr(origin_pandas,duplicate_pandas,origin_folder_path,**walk_folder_args):
    ocr_engines = walk_folder_args['ocr_engines']
    temp_folder_path = walk_folder_args['temp_folder_path']
    prepare_engine = walk_folder_args['engine_switch']
    result_pandas = origin_pandas
    # 获取文件夹内所有的jpg和pdf文件个数
    cnt_file = len({p.resolve() for p in Path(origin_folder_path).glob("*") if p.suffix in [".jpg", ".pdf"]})
    # 如果要包括子目录中的文件，则为：
    # cnt_total = len({p.resolve() for p in Path(origin_folder_path).glob("**/*") if p.suffix in [".jpg", ".pdf"]})
    inv_dict = {}  #发票字典初始化  
    #从origin_pandas 构建inv_dict字典(票号:文件路径)
    if not result_pandas.empty:
        for i, (index, row) in enumerate(result_pandas.iterrows()):
            if row['01票号'] is np.NAN: #如果票号是空，则跳过
                continue
            if row['01票号'] not in inv_dict:
                inv_dict[row['01票号']] = [row['file_path']]
            else:
                inv_dict[row['01票号']].append(row['file_path'])
    if not duplicate_pandas.empty:
        for i, (index, row) in enumerate(duplicate_pandas.iterrows()):
            if row['重复票号'] is np.NAN: #如果票号是空，则跳过
                continue
            if row['重复票号'] not in inv_dict:
                inv_dict[row['重复票号']] = [row['file_path']]
            else:
                inv_dict[row['重复票号']].append(row['file_path'])   
    cnt_done = 0
    cnt_duplicate = 0
    if not origin_pandas.empty:
        cnt_done = len(origin_pandas.loc[origin_pandas['file_path'].notnull(),:])
    if not duplicate_pandas.empty:
        cnt_duplicate = len(duplicate_pandas.loc[duplicate_pandas['file_path'].notnull(),:])
    for file_name in os.listdir(origin_folder_path): #只在本层文件夹内遍历
        file_path = os.path.join(origin_folder_path, file_name)
        if os.path.isfile(file_path): #排除file_name是文件夹的情况
            pr,nm,fr,ex = pathsplit(file_path)
            if ex not in ['.pdf','.jpg']:
                continue

            inv_out_of_result_pandas = True
            inv_out_of_duplicate_pandas = True
            # 在上次结果文件和重复文件记录中查找文件路径：
            try:
               inv_out_of_result_pandas = result_pandas.loc[result_pandas['file_path']==file_path,:].empty
               inv_out_of_duplicate_pandas = duplicate_pandas.loc[duplicate_pandas['file_path']==file_path,:].empty
            except:
                pass
            #如果文件路径在上次结果文件和重复文件记录中查询结果不为空，即曾识别过，则跳过该文件
            if not(inv_out_of_result_pandas and inv_out_of_duplicate_pandas):
                continue        
            result_series_orderdic = OrderedDict() #定义series有序字典
            err_info = '' #错误记录初始化
            if ex == '.pdf':
                inv_code = ''
                pdf_trans_file_fr = fr
                pdf_trans_file_ex = '.xlsx'
                # pdf_trans_file_ex = '.txt'
                pdf_trans_file_nm = pdf_trans_file_fr + pdf_trans_file_ex
                pdf_trans_folder_name = 'temp_pdf_trans_excel'
                pdf_trans_folder_path = os.path.join(temp_folder_path, pdf_trans_folder_name)
                
                if not os.path.exists(pdf_trans_folder_path):
                    os.mkdir(pdf_trans_folder_path)

                pdf_trans_file_path = os.path.join(pdf_trans_folder_path, pdf_trans_file_nm)
                
                if not os.path.exists(pdf_trans_file_path):
                    trans_type = '.xlsx'
                    # trans_type = '.txt'
                    pdf_trans_file_path = Pdf_tans_to(file_path, pdf_trans_file_path, trans_type = trans_type, temp_pdf_trans_excel_out = True)
                
                if os.path.exists(pdf_trans_file_path):
                    result_series_orderdic, err_info, inv_dict = Tele_inv_ocr(ocr_engines, result_series_orderdic, inv_dict, file_path, pdf_trans_file_path, err_info, engine_switch = precise_engine)   

                if len(result_series_orderdic) != 0:
                    if '01票号' in result_series_orderdic:
                        inv_code = result_series_orderdic['01票号'][0].values[0]
                        #票号添加到票号字典
                        if inv_code not in inv_dict:
                            inv_dict[inv_code] = [file_path]
                        else:
                            if file_path not in inv_dict[inv_code]:
                                inv_dict[inv_code].append(file_path)
                        if len(inv_dict[inv_code]) > 1: #如果该票号的发票重复，跳出本张图片循环
                            if duplicate_pandas.empty:
                                duplicate_pandas = pd.DataFrame(data={'重复票号':[inv_code],'file_path':[file_path]}) 
                            else:
                                duplicate_pandas = pd.concat([duplicate_pandas, pd.DataFrame(data={'重复票号':[inv_code],'file_path':[file_path]})], ignore_index = True, axis = 0)    
                            Log_result_file(duplicate_pandas,result_file_path,duplicate_sheet_name)
                            cnt_duplicate = cnt_duplicate + 1
                            print(datetime.now().strftime("%H:%M:%S"),file_path, 'Skip. ','\n\t\tDuplicate:', inv_code,inv_dict[inv_code][0])
                            #发票号重复，跳出本次识别
                            continue 
                else:
                    #如果没有结果，转成图片识别
                    pdf_trans_file_ex = '.jpg'
                    pdf_trans_file_nm = pdf_trans_file_fr + '.jpg'
                    pdf_trans_folder_name = 'temp_pdf_trans_jpg'
                    pdf_trans_folder_path = os.path.join(temp_folder_path, pdf_trans_folder_name)
                    pdf_trans_jpg_file_path = os.path.join(pdf_trans_folder_path, pdf_trans_file_nm)
                    pdf_trans_jpg_file_path = Pdf_tans_jpg(file_path, pdf_trans_jpg_file_path, temp_pdf_trans_jpg_out = True)

                    if len(pdf_trans_jpg_file_path)>0:
                        if os.path.exists(pdf_trans_jpg_file_path):
                            #如果传回了转成图片的路径，并且路径存在，读取jpg路径，付给file_path,转成ocr识别：
                            print('\n\nPDF转成图片识别：',pdf_trans_jpg_file_path,'【此模块待添加。】\n\n')
                        

            elif str.lower(ex) == '.jpg':        
                known_dict = {} #初始化
                inv_code ='' #初始化
                temp_img_trans_excel_folder = os.path.join(temp_folder_path,'temp_img_trans_excel')
                img_trans_xls_name = 'result_' + fr +  '.xlsx' 
                img_trans_xls_path = os.path.join(temp_img_trans_excel_folder, img_trans_xls_name)

                if os.path.exists(img_trans_xls_path):
                    origin_df = pd.read_excel(img_trans_xls_path, sheet_name=0,header=0,index_col=0,na_values=None, keep_default_na=False, dtype=object) #读取表格
                else:
                    known_dict = Crop_known_from_qrcode(file_path)
                    if len(known_dict)>0:
                        inv_code = known_dict['01票号'].values[0]
                        #票号添加到票号字典
                        if inv_code not in inv_dict:
                            inv_dict[inv_code] = [file_path]
                        else:
                            if file_path not in inv_dict[inv_code]:
                                inv_dict[inv_code].append(file_path)
                        if len(inv_dict[inv_code]) > 1: #如果该票号的发票重复，跳出本张图片循环
                            if duplicate_pandas.empty:
                                duplicate_pandas = pd.DataFrame(data={'重复票号':[inv_code],'file_path':[file_path]}) 
                            else:
                                duplicate_pandas = pd.concat([duplicate_pandas, pd.DataFrame(data={'重复票号':[inv_code],'file_path':[file_path]})], ignore_index = True, axis = 0)    
                            Log_result_file(duplicate_pandas,result_file_path,duplicate_sheet_name)
                            cnt_duplicate = cnt_duplicate + 1
                            print(datetime.now().strftime("%H:%M:%S"),file_path, 'Skip. ','\n\t\tDuplicate:', inv_code,inv_dict[inv_code][0])
                            #发票号重复，跳出本次识别
                            continue 
                    origin_df = Ocr_func(ocr_engines, img_path = file_path, temp_folder_path = temp_folder_path, 
                        range_title = '', known_dict=known_dict, ocr_excel_out = ocr_excel_out, draw_result_out = draw_result_out, engine_switch=prepare_engine)  #识别为原始文本df
                if not origin_df.empty:
                    result_series_orderdic, err_info = Loc_range_content_pandas(ocr_engines, origin_df, result_series_orderdic, err_info, known_dict, file_path, temp_folder_path, enhance = enhance, engine_switch=precise_engine) #处理为结果series字典
                    if len(result_series_orderdic['01票号']) > 0:
                        inv_code = result_series_orderdic['01票号'].values[0]
                        # assert isinstance(inv_code,str)
                        # assert len(inv_code) == 8 or len(inv_code) == 20
                        if inv_code not in inv_dict:
                            inv_dict[inv_code] = [file_path]
                        else:
                            if file_path not in inv_dict[inv_code]:
                                inv_dict[inv_code].append(file_path)
            if len(inv_code)>0 and inv_code in inv_dict and len(inv_dict[inv_code]) >1:
            # duplicate_df = pd.read_excel(result_file_path, sheet_name=duplicate_sheet_name,index_col=0,header = 0,keep_default_na=True,dtype=object) #读取表格
                if duplicate_pandas.empty:
                    duplicate_pandas = pd.DataFrame(data={'重复票号':[inv_code],'file_path':[file_path]}) 
                else:
                    duplicate_pandas = pd.concat([duplicate_pandas, pd.DataFrame(data={'重复票号':[inv_code],'file_path':[file_path]})], ignore_index = True, axis = 0)    
                Log_result_file(duplicate_pandas,result_file_path,duplicate_sheet_name)
                cnt_duplicate = cnt_duplicate + 1
                print(datetime.now().strftime("%H:%M:%S"),file_path, 'Skip. ','\n\t\tDuplicate:', inv_code,inv_dict[inv_code][0])
                continue #如果发票号不只一张，跳出本次识别
            #series列表合成dataframe:
            bind_df = pd.DataFrame([result_series_orderdic[series_title][0] if isinstance(result_series_orderdic[series_title], list) else result_series_orderdic[series_title] for series_title in result_series_orderdic]).T
            columns_list =  ['01票号','02代码','03日期','04购方','05购方税号','06品名','07单位','08数量','09单价','10税前',
                    '11税率','12税额','13合计税前','14合计税额','15总额','16大写','17销方','18销方税号'] 
            if len(bind_df) == 0:
                bind_df = pd.DataFrame(columns = columns_list)
            result_df = bind_df.copy() #浅拷贝，防止下面填充提示错误
            result_df['file_path'] = ''
            if len(result_df) == 0:
                result_df = result_df.append({'file_path':file_path},ignore_index = True) #追加文件路径到第一行
            else:
                result_df['file_path'].values[0] = file_path #追加文件路径到第一行
            result_df['err_info'] = ''
            result_df.loc[result_df.index[0],'err_info'] = err_info #追加错误提示到第一行
            # 填充处理：务必先处理na值，再进行后续处理。
            
            result_df = Fill_na_result(result_df)

            if result_pandas.empty:
                result_pandas = result_df
            else:
                result_pandas = pd.concat([result_pandas, result_df], ignore_index = True, axis = 0)

            result_pandas = Check_result(result_pandas) #检查和修改结果 每识别一个文件，重新检查前面所有的发票
            #每识别一个文件，写入结果文件，防止中间出错导致未保存结果而重复识别，以实现断点接续，提高总体的效率：
            Log_result_file(result_pandas,result_file_path,result_sheet_name)
            # writer = pd.ExcelWriter(result_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
            # duplicate_pandas.to_excel(writer,sheet_name=duplicate_sheet_name)
            # writer.close()

            #-----添加文件路径超链接------
            Add_hyperlink(result_file_path,result_sheet_name)

            cnt_done = cnt_done + 1
            print(datetime.now().strftime("%H:%M:%S"),file_name, inv_code,'done: ' + str(cnt_done) + ' / ' + str(cnt_file))
    # cnt_dict = {'cnt_file':cnt_file,'cnt_done':cnt_file,'cnt_done':cnt_duplicate}
    return result_pandas,duplicate_pandas

#-------------------------func: Ocr_func()--------------------------------
                    
# ocr image to origin_DataFrame. 
@nb.autojit
def Ocr_func(ocr_engines, img_path, temp_folder_path,  range_title='', known_dict = {}, ocr_excel_out = True, draw_result_out = False, engine_switch = 0) ->object: #DataFrame            

    p,n,fr,ex = pathsplit(img_path) #拆分路径

    temp_img_trans_excel_folder = os.path.join(temp_folder_path,'temp_img_trans_excel')
    temp_draw_result_folder = os.path.join(temp_folder_path,'temp_draw_result')
    if engine_switch == 0:
        engine = 'mb'
    elif engine_switch == 1:
        engine = 'pp'
    elif engine_switch == 2:
        engine = 'sv'
    if range_title =='':
        img_trans_xls_name = 'result(' + engine + ')_' + fr + '.xlsx' 
    else:
        img_trans_xls_name = 'result(' + engine + ')_' + fr + '_' + range_title + '.xlsx' 
    img_trans_xls_path = os.path.join(temp_img_trans_excel_folder, img_trans_xls_name)

    if not os.path.exists(temp_img_trans_excel_folder):
        Create_clear_dir(temp_img_trans_excel_folder)
    if not os.path.exists(temp_draw_result_folder):
        Create_clear_dir(temp_draw_result_folder)

    result = '' #结果初始化

    if engine_switch == 1:
        paddleOcr = ocr_engines[engine_switch] 
        results = paddleOcr.ocr(img_path, cls=True)  #识别图像----------------
        df0 = pd.DataFrame(data=results,columns=['pix','result'])
        df1 = pd.concat([pd.DataFrame(df0['pix'].values.tolist(),columns=['lu','ru','rd','ld']), pd.DataFrame(df0['result'].values.tolist(),columns=['content','trust'])], axis=1)
        title_list = ['lu', 'ru', 'rd', 'ld']
        df = df1[['content','trust']]
        for i, title in enumerate(title_list):
            df = pd.concat([df, pd.DataFrame(df1[title].values.tolist(), columns=[title + 'w', title + 'h'])], axis=1)

        if ocr_excel_out == True:
            df.to_excel(img_trans_xls_path, index=False)

        if draw_result_out == True:
            # draw result
            from PIL import Image
            image = Image.open(img_path).convert('RGB')
            boxes = [line[0] for line in result]
            txts = [line[1][0] for line in result]
            scores = [line[1][1] for line in result]
            im_show = draw_ocr(image, boxes, txts, scores, font_path='./fonts/simfang.ttf')
            im_show = Image.fromarray(im_show)
            if range_title =='':
                draw_result_name = 'draw_result_' + fr + ex
            else:
                draw_result_name = 'draw_result_' + fr + '_' + range_title + ex 
            draw_result_path = os.path.join(temp_draw_result_folder, draw_result_name)
            im_show.save(draw_result_path)
        
    elif engine_switch == 0 or engine_switch == 2:
        hubOcr = ocr_engines[engine_switch]
        img = cv_imread(img_path)
        np_images = [img]
#         np_images = [cv2.imdecode(np.fromfile(jpgfile, dtype=np.uint8), cv2.IMREAD_COLOR)]
        #---------使用识别引擎：
        hub_result = hubOcr.recognize_text(
            images=np_images,  # 图片数据，ndarray.shape 为 [H, W, C]，BGR格式
            use_gpu=False,  # 是否使用 GPU。否即False,是即请先设置CUDA_VISIBLE_DEVICES环境变量
            output_dir=temp_draw_result_folder,  # 图片的保存路径
            visualization=True,  # 是否将识别结果保存为图片文件
            box_thresh=0.5,  # 检测文本框置信度的阈值
            text_thresh=0.5)  # 识别中文文本置信度的阈值
        results = hub_result[0]['data']
        df = pd.DataFrame()
        column_list = ['content','confdence','luw','luh','ruw','ruh','rdw','rdh','ldw','ldh']
        for infomation in results:
            content = infomation['text']
            confidence = infomation['confidence']
            box = infomation['text_box_position']
            luw,luh,ruw,ruh = box[0][0],box[0][1],box[1][0],box[1][1]
            rdw,rdh,ldw,ldh = box[2][0],box[2][1],box[3][0],box[3][1]
            line = [content,confidence,luw,luh,ruw,ruh,rdw,rdh,ldw,ldh]
            line_df = pd.DataFrame(data = line,index = column_list).T
            if df.empty:
                df = line_df
            else:
                df = pd.concat([df, line_df], axis=0, ignore_index=True)
        if ocr_excel_out == True:
            df.to_excel(img_trans_xls_path, index = False)
    return df

#----------------------func: Loc_range_content_pandas()----------------------------
# 根据预识别表定位数据
@nb.autojit
def Loc_range_content_pandas(ocr_engines, df, result_series_orderdic, err_info, known_dict, file_path, temp_folder_path, enhance=False, engine_switch=0): #DataFrame

    user_name, user_code  = '', ''

    df['content'].astype(str) #content列转换为字符格式
    
    #计算各识别区域中心点的w和h
    df['center_w']=(df.loc[:,'luw']+df.loc[:,'rdw'])/2
    df['center_h']=(df.loc[:,'luh']+df.loc[:,'rdh'])/2

    # w_prem_title_distance = 1605  #300像素表中，开票日期左上与价税合计左上w距离
    # h_prem_title_distance = -907  #300像素表中，开票日期左上与价税合计左上h距离

    #以下代码注销，w_this_distance h_this_distance w_ratio,h_ratio 都使用固定值 allen 2022.11.16
    # #预先测量标题变量列表
    # prem_title_list = ['价税合计','开票日期']
    # temp_df_prem_title0 = df.loc[df['content'].str.contains(prem_title_list[0])]
    # temp_df_prem_title1 = df.loc[df['content'].str.contains(prem_title_list[1])]
    
    # if temp_df_prem_title0.empty or temp_df_prem_title1.empty:
    #     #如果没找到比例定位字段
    #     err_zero = 'Serious fault: title[' + ' or '.join(prem_title_list) + '] not found!!!'
    #     w_this_distance = 1600
    #     h_this_distance 
    #     w_ratio = 1
    #     h_ratio = 1
    # else:
    #     # loc0:预测量坐标_价税合计位置，loc1:预测量坐标_开票日期位置，loc_zero:汉字金额原点位置
    #     [w_this_loc_tradtitle,h_this_loc_tradtitle] = temp_df_prem_title0.loc[:,['luw','luh']].values[0]
    #     [w_this_loc_datetitle,h_this_loc_datetitle] = temp_df_prem_title1.loc[:,['luw','luh']].values[0]
    #     w_this_distance = w_this_loc_datetitle - w_this_loc_tradtitle
    #     h_this_distance = h_this_loc_datetitle - h_this_loc_tradtitle
    #     w_ratio = round(w_this_distance/w_prem_title_distance,2)
    #     h_ratio = round(h_this_distance/h_prem_title_distance,2)

    # w_this_distance = w_prem_title_distance
    # h_this_distance = h_prem_title_distance
    w_ratio = 1
    h_ratio = 1

    w_this_loc_tradtitle = 240
    h_this_loc_tradtitle = 1170
        
    # 用价税合计标题字段坐推算汉字金额区域坐标范围，并查找确定汉字金额的坐标 
    # tolerance = 18 #容错像素范围
    # 汉字金额区域左右上下坐标与价税合计标题距离:
    min_w_zero_distance, max_w_zero_distance ,min_h_zero_distance, max_h_zero_distance \
        = 521,1550,-33,98
    # 具体某张发票中的原点（即汉字金额）的宽度和长度区块：
    min_w_zero = w_this_loc_tradtitle + w_ratio * min_w_zero_distance
    max_w_zero = w_this_loc_tradtitle + w_ratio * max_w_zero_distance
    min_h_zero = h_this_loc_tradtitle + h_ratio * min_h_zero_distance
    max_h_zero = h_this_loc_tradtitle + h_ratio * max_h_zero_distance
    loc_trad_range = [min_w_zero, max_w_zero, min_h_zero, max_h_zero]
    
    # 查询原点字段的条件：
    cond_trad = [
        '16大写',
        'direct', #原点后面也通过裁切重新识别。（解决直接获取出现的识别汉字不全的情况）
        'contains', 
        '[圆角分整零壹贰叁肆伍陆柒捌玖拾佰仟万亿]{2,}', 
        '1', #提取数量限制
        {'direct':loc_trad_range}
        ]

    known_dict = Loc_jpg_content(df, cond_trad, order_dict=known_dict)
    if len(known_dict['16大写'][1]) > 0:
        (w_zero, h_zero) = known_dict['16大写'][1]
    else:
        err_info = err_info + '识别失败！未找到大写金额内容。'
        #设定一个默认坐标
        w_zero = 750
        h_zero = 1180

    # 其他字段区块查询条件：
    range_list = [
        [
            #发票号码
            '01票号',  #0 区域代号
            ['known','crop'], #1 前面先qrcode，通过字典的known获取，其次crop识别
            'extract', #2 提取方式
            '^\D*(\d{8})$',   #3 正则表达式
            '1', #4 提取数量限制
            { #5
            'crop':
                [int(w_zero + w_ratio *  (1430)), 
                int(w_zero + w_ratio *  (1685)), 
                int(h_zero + h_ratio * (-990)), 
                int(h_zero + h_ratio * (-900))], 
            'known':
                known_dict
            }       
        ], 
        [
            #发票代码
            '02代码',  #字段标题
            ['known','crop'], #1 前面先qrcode，通过字典的known获取，其次crop识别
            'extract', 
            '([a-zA-Z0-9]{10})$', 
            '1', #提取数量限制
            {
            'crop':
                [int(w_zero + w_ratio *  (-475)), 
                int(w_zero + w_ratio *  (80)), 
                int(h_zero + h_ratio * (-1100)), 
                int(h_zero + h_ratio * (-920))]
            } 
        ],
        [
            #开票日期
            '03日期',
            ['known','crop'], #数据提取方式列表：汉字金额和前面先qrcode的，通过字典的known获取，其次crop、direct
            'extract', 
            '(\d{4}\s*年\s*\d{2}\s*月\s*\d{2}\s*日)$',
            '1', #提取数量限制
            {
            'direct':
                [int(w_zero + w_ratio *  (1100)), 
                int(w_zero + w_ratio *  (1637)), 
                int(h_zero + h_ratio * (-925)), 
                int(h_zero + h_ratio * (-840))],
            'crop':
                [int(w_zero + w_ratio *  (1300)), 
                int(w_zero + w_ratio *  (1637)), 
                int(h_zero + h_ratio * (-925)), 
                int(h_zero + h_ratio * (-840))],
            }
        ],
        [
            #买方名称
            '04购方',
            ['crop'],
            'extract',
            '([\(\)（）\u4e00-\u9fa5]{8,30})', 
            '1', 
            {
            'crop':
                [int(w_zero + w_ratio *  (-320)), 
                int(w_zero + w_ratio *  (600)), 
                int(h_zero + h_ratio * (-800)), 
                int(h_zero + h_ratio * (-680))],
            }
        ],
        [
            #买方税号
            '05购方税号',
            ['direct'],
            'extract', 
            '([a-zA-Z0-9]{18})$', 
            '1', 
            {
            'direct':
                [int(w_zero + w_ratio *  (-240)), 
                int(w_zero + w_ratio *  (540)), 
                int(h_zero + h_ratio * (-800)), 
                int(h_zero + h_ratio * (-680))],
            'crop':
                [int(w_zero + w_ratio *  (-320)), 
                int(w_zero + w_ratio *  (600)), 
                int(h_zero + h_ratio * (-800)), 
                int(h_zero + h_ratio * (-680))],
            }
        ],
        [
            #商品名称  ok
            '06品名',
            ['crop'],
            'contains', 
            '^[\*冰水米\+]?(\S*[制品]\S*[\*冰水米\+]?\S+)$', 
            'n', 
            {
            'crop':
                [int(w_zero + w_ratio *  (-670)), 
                int(w_zero + w_ratio *  (640)), 
                int(h_zero + h_ratio * (-560)), 
                int(h_zero + h_ratio * (-100))],
            #在crop图中的位置限定
            'center_limit': [10, 500, 10, 450],
            }
        ],
        [
            #单位
            '07单位',
            ['crop'],
            'contains', 
            '^\D{1,8}$', 
            'n', 
            {
            'crop': #和品名同一区块总体聚类后再识别，通过位置提取，降低了漏识率
                [int(w_zero + w_ratio *  (-670)), 
                int(w_zero + w_ratio *  (640)), 
                int(h_zero + h_ratio * (-560)), 
                int(h_zero + h_ratio * (-100))],
            #在crop图中的位置限定:
            'center_limit': [820,1100,10,450]
            }    
        ],
        [
            #数量
            '08数量',
            ['crop'],
            'contains', 
            '^\d+$|^\d+\.\d+$',
            'n', 
            {
            'crop':
                [int(w_zero + w_ratio *  (440)), 
                int(w_zero + w_ratio *  (640)), 
                int(h_zero + h_ratio * (-510)), 
                int(h_zero + h_ratio * (-100))],
            }    
        ],
        [
            #单价
            '09单价',
            ['crop'],
            'contains', 
            '^[\.:：]?\d+[\.:：]?\s*\d*\s*$', 
            'n', 
            {
            'crop':
                [int(w_zero + w_ratio *  (635)), 
                int(w_zero + w_ratio *  (890)), 
                int(h_zero + h_ratio * (-510)), 
                int(h_zero + h_ratio * (-100))],
            }
        ],
        [
            #商品明细税前金额
            '10税前',
            ['crop'],
            'contains', 
            '^\s*[+-]?(?:\d+|\d{1,3}(?:,\d{3})*)[\.:：]\s*\d{2}\s*$', 
            'n', 
            {
            'crop':
                [int(w_zero + w_ratio *  (980)), 
                int(w_zero + w_ratio *  (1240)), 
                int(h_zero + h_ratio * (-510)), 
                int(h_zero + h_ratio * (-100))],
            }
        ],
        [
            #税率
            '11税率',
            ['crop'],
            'contains', 
            '^\d{1,2}\s*%$', 
            '1', 
            {
            'crop':
                [int(w_zero + w_ratio *  (1240)), 
                int(w_zero + w_ratio *  (1350)), 
                int(h_zero + h_ratio * (-510)), 
                int(h_zero + h_ratio * (-100))],
            }
        ],
        [
            #商品明细税额
            '12税额',
            ['crop'],
            'contains', 
            '^\s*[+-]?(?:\d+|\d{1,3}(?:,\d{3}))[\.:：]?\s*\d{0,2}\s*\D*', 
            'n', 
            {
            'crop':
                [int(w_zero + w_ratio *  (1380)), 
                int(w_zero + w_ratio *  (1700)), 
                int(h_zero + h_ratio * (-510)), 
                int(h_zero + h_ratio * (-100))],
            }    
        ],
        [
            #合计税前金额
            '13合计税前',
            ['known','crop'], #1 前面先qrcode，通过字典的known获取，其次识别
            'contains', 
            '[¥￥]?s*[+-]?(?:\d+|\d{1,3}(?:,\d{3})*)[\.:：]\s*\d{2}\s*$',
            '1', 
            {
            'crop':
                [int(w_zero + w_ratio *  (880)), 
                int(w_zero + w_ratio *  (1235)), 
                int(h_zero + h_ratio * (-100)), 
                int(h_zero + h_ratio * (-10))],
            'known':
                known_dict
            }
        ],
        [
            #合计税额
            '14合计税额',
            ['crop'],
            'contains', 
            '[¥￥]?s*[+-]?(?:\d+|\d{1,3}(?:,\d{3})*)[\.:：]?\s*\d{0,2}\s*$',
            '1', 
            {
            'crop':
                [int(w_zero + w_ratio *  (1300)), 
                int(w_zero + w_ratio *  (1710)), 
                int(h_zero + h_ratio * (-110)), 
                int(h_zero + h_ratio * (0))],
            }
        ],
        [
            #合计总额小写
            '15总额',
            ['crop'],
            'contains', 
            '[¥￥]?s*[+-]?(?:\d+|\d{1,3}(?:,\d{3})*)[\.:：]\s*\d{2}\s*$',
            '1', 
            {
            'crop':
                [int(w_zero + w_ratio *  (1220)), 
                int(w_zero + w_ratio *  (1700)), 
                int(h_zero + h_ratio * (-20)), 
                int(h_zero + h_ratio * (70))],
            }   
        ],   
        [
            #合计大写
            '16大写',
            ['known'],
            known_dict
        ],
        [
            #销方名称
            '17销方',
            ['crop'],
            'extract', 
            # '([\(\)（）\u4e00-\u9fa5]{8,30})', 
            '([\(\)（）\u4e00-\u9fa5]{8,30}[办|处|公|司|厂|社|部])$', 
            '1', 
            {
            'crop':
                [int(w_zero + w_ratio *  (-280)), 
                int(w_zero + w_ratio *  (540)), 
                int(h_zero + h_ratio * (60)), 
                int(h_zero + h_ratio * (165))],
            }
        ],
        [
            #销方税号
            '18销方税号',
            ['direct'],
            'extract',
            '([a-zA-Z0-9]{18})$', 
            '1', 
            {
            'direct':
                [int(w_zero + w_ratio *  (-260)), 
                int(w_zero + w_ratio *  (600)), 
                int(h_zero + h_ratio * (100)), 
                int(h_zero + h_ratio * (220))],
            'crop':
                [int(w_zero + w_ratio *  (-320)), 
                int(w_zero + w_ratio *  (600)), 
                int(h_zero + h_ratio * (100)), 
                int(h_zero + h_ratio * (220))],
            }
        ]
    ]

    # cv2获取图像对象
    img_inv = cv_imread(file_path)  #获取需要crop识别前的整张发票原始图片

    err_info = '' #初始化此发票错误提示信息
    set_h_adjust = 0 #初始化crop的高度像素微调像素
    # 开始遍历字段条件列表逐个提取字段：
    for i, cond_list in enumerate(range_list):
        range_title = cond_list[0]
        loc_method = cond_list[1]
        result_series_orderdic[range_title] = pd.Series() #返回值初始化
        
        if 'known' in loc_method:
            if range_title in known_dict:
                known = True
                result_series_orderdic[range_title] = known_dict[range_title] #追加已知字典中的字段识别信息到字典 
                if len(result_series_orderdic[range_title]) > 0:
                    continue  #如果已赋值，跳出本轮循环,不再执行本轮后面的语句
        if 'crop' in loc_method:
            crop_folder_name = 'crop'
            crop_folder_path = os.path.join(temp_folder_path, crop_folder_name) 
            if not os.path.exists(crop_folder_path):
                Create_clear_dir(crop_folder_path)
            result_series_orderdic, get_h_adjust = Crop_ocr(ocr_engines, result_series_orderdic,known_dict, img_inv, file_path, crop_folder_path, set_h_adjust, cond_list, enhance, engine_switch = engine_switch)
            if range_title == '01票号':
                #用识别01票号获取的高度像素调整结果如果大于5，设定为其他区域裁切高度的调整值。识别其他区域返回的系数不用管。
                if get_h_adjust > 5:
                    set_h_adjust = get_h_adjust
                if len(result_series_orderdic[range_title]) > 0:
                    continue  #如果已赋值，跳出本轮循环,不再执行本轮后面的语句
        if 'direct' in loc_method:
            result_series_orderdic = Loc_jpg_content(df, cond_list, order_dict=result_series_orderdic)
    
    return result_series_orderdic, err_info

#------------------Crop_known_from_qrcode Func()------------------
def Crop_known_from_qrcode(file_path) ->dict:

    known_dict = {} #返回值初始化
    pr,nm,fr,ex = pathsplit(file_path)
    qrcode_folder_name = 'temp_crop_qrcode'
    qrcode_folder_path = os.path.join(temp_folder_path, qrcode_folder_name)
    if not os.path.exists(qrcode_folder_path):
        Create_clear_dir(qrcode_folder_path)
    qrcode_file_name = 'qrcode_' + nm
    qrcode_file_path = os.path.join(qrcode_folder_path, qrcode_file_name)
    qrcode_image_crop = Crop_qrcode_image(file_path, qrcode_file_path)  # -----------切割处理二维码图片

    qrcode_result = ''
    if qrcode_image_crop == True: #如果二维码切图返回为True
        qrcode_result = qrcode_recongnize(qrcode_file_path)    #------------二维码识别
    if len(qrcode_result) > 0:
        if len(qrcode_result) > 20:
            qrcode_list = qrcode_result.split(',') 
            for index, range_title in enumerate(['02代码','01票号','13合计税前','04日期']): #二维码各字段结果逐个赋值给knowndict
                known_dict[range_title] = pd.Series(data=qrcode_list[index+2],name = range_title)
    
    return known_dict


# ------------------Crop_qrcode_image()----------------------
#切割二维码图片并放大像素
def Crop_qrcode_image(origin_file_path,crop_file_path):
    # 切割二维码图片
    result = False #结果初始化
    img_inv = cv_imread(origin_file_path)
    img_crop = img_inv[100:400, 50:350]  # h, w
    img_magnify = cv2.resize(img_crop, (1200, 1200))
    cv2.imencode('.jpg', img_magnify)[1].tofile(crop_file_path) 
    if os.path.exists(crop_file_path):
        result = True
    return result


# ------------------qrcode_recongnize()----------------------
# 二维码识别：
def qrcode_recongnize(file_path, method = 'cv2', drawframe = False, enhance=False): #method：pil or cv2
    pr = os.path.split(file_path)[0]
    nm = os.path.split(file_path)[1]
    output_img_path = os.path.join(pr, 'draw_qrcode_' + nm)
    
    #方式一：cv2 方式
    if method =='cv2':
        img = cv_imread(file_path)
        gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        barcodes =pyzbar.decode(gray_img)
    #     print(barcodes)
        barcodeData = ''
        if len(barcodes) >0 :
            for barcode in barcodes:
                # 提取条形码的边界框的位置
                # 画出图像中条形码的边界框
                (x, y, w, h) = barcode.rect
                cv2.rectangle(img, (x, y), (x + w, y + h), (255, 255, 0), 2)
                # 条形码数据为字节对象，所以如果我们想在输出图像上
                #  画出来，就需要先将它转换成字符串
                barcodeData = barcode.data.decode("utf-8")
                if len(barcodeData) > 20:
                    if drawframe == True:
                        from PIL import Image, ImageFont, ImageDraw
                        # 绘出图像上条形码的数据和条形码类型
                        barcodeType = barcode.type
                        # 把cv2格式的图片转成PIL格式的图片然后在上标注二维码和条形码的内容
                        img_PIL = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
                        # 参数（字体，默认大小）
                        font = ImageFont.truetype('STFANGSO.TTF', 25)
                        # 字体颜色
                        fillColor = (0,255,0)
                        # 文字输出位置
                        position = (x, y-25)
                        # 输出内容
                        strl = barcodeData
                        # 需要先把输出的中文字符转换成Unicode编码形式(str.decode("utf-8))  
                        # 创建画笔
                        draw = ImageDraw.Draw(img_PIL)
                        draw.text(position, strl, font=font,fill=fillColor)
                        # 使用PIL中的save方法保存图片到本地
                        img_PIL.save(output_img_path, 'jpeg')
                        # 向终端打印条形码数据和条形码类型
                        # print("扫描结果==》 类别： {0} 内容： {1}".format(barcodeType, barcodeData))
                    break

        return barcodeData
        
    elif method == 'pil':
        #方式二：pil+qrcode
        from PIL import Image, ImageEnhance
        img = Image.open(file_path).convert('RGB')
        if enhance == True:
            # 增加亮度
            img = ImageEnhance.Brightness(img).enhance(1.0)
            # 锐利化
            img = ImageEnhance.Sharpness(img).enhance(1.5)
            # 增加对比度
            img = ImageEnhance.Contrast(img).enhance(2.0)
            # 灰度化
            img = img.convert('L')
        # 解码二维码
        decoded = pyzbar.decode(img)
        result = decoded[0][0].decode('utf-8')
        return result


# ------------------------------func:Crop_ocr-----------------------
# 功能：切割图片识别
def Crop_ocr(ocr_engines, result_series_orderdic, known_dict,img_inv, file_path, crop_folder_path, set_h_adjust, cond_list, enhance = False, engine_switch = 0):
    pr,nm,fr,ex = pathsplit(file_path)
    range_title = cond_list[0]
    loc_method = cond_list[1]
    reg_type = cond_list[2]
    reg = cond_list[3]
    count_limit = cond_list[4]
    loc_dict = cond_list[5]
    chop_pix = loc_dict['crop']
    [min_w,max_w,min_h,max_h] = chop_pix
    adjust_ratio_dict = {'02代码':1, '03日期':1,'10税前':0.6,'11税率':0.7,'12税额':0.8}
    if range_title in adjust_ratio_dict:
        adjust_ratio = adjust_ratio_dict[range_title]
        min_h = min_h - int(set_h_adjust * adjust_ratio) #用微调系数和条件字段调节比例对裁切高度做微调，得到截取的新区域坐标
        max_h = max_h - int(set_h_adjust * adjust_ratio) 
    crop_center_h = (max_h - min_h)//2 #获取截取后的区域的中心高度，即原图截取坐标高度差的一半
    img_crop = img_inv[min_h:max_h, min_w:max_w]
    enhance_title = ['04购方','05购方税号','06品名','07单位','16大写','17销方','18销方税号']
    if enhance == True:
        if range_title in enhance_title:
            img_pil = cv2_pil(img_crop)
            img_enhance = pil_enhance(img_pil)
            img_crop = pil_cv2(img_enhance)
    crop_file_name = 'crop_'+ range_title + '_' +nm
    crop_file_path = os.path.join(crop_folder_path, crop_file_name)
    cv2.imencode('.jpg', img_crop)[1].tofile(crop_file_path) 
    df = Ocr_func(ocr_engines, img_path = crop_file_path, temp_folder_path = crop_folder_path, 
                range_title = range_title, known_dict=known_dict,ocr_excel_out = True, draw_result_out = True, engine_switch = engine_switch)
    get_h_adjust = 0 #高度微调参数赋初始值
    result_sr = pd.Series(name = range_title)  #结果初始化
    # if range_title in ['09单价','02代码','03日期']:  #调试
    #     print(range_title)
    if reg_type == 'extract':
        cond_df = df['content'].str.extract(reg)
        cond_df.loc[:,['luh','ldh']] = df.loc[:,['luh','ldh']]
        content_result = pd.notna(cond_df[0])
        if 'center_limit' in loc_dict: #如果字典有中心位置限制条件数据，则核对数据中心点位置是否符合条件
            center_df = df #赋值给中间临时表center_df，用以计算中心位置是否满足限制条件
            center_df[['luw','ruw','luh','ldh']].astype(int)
            center_df['center_w'] = (center_df['luw'] + center_df['ruw']) //2 
            center_df['center_h'] = (center_df['luh'] + center_df['ldh']) //2
            [center_w_min, center_w_max, center_h_min, center_h_max] = loc_dict['center_limit']
            cond_center = (center_w_min <= center_df.loc[:,'center_w']) & (center_df.loc[:,'center_w'] <= center_w_max) & \
                (center_h_min <= center_df.loc[:,'center_h']) & (center_df.loc[:,'center_h'] <= center_h_max) #坐标限定 
            content_result = content_result & cond_center
        temp_df = df.loc[cond_df[content_result].index,:]
        if not temp_df.empty:
            temp_sr = temp_df.iloc[:,0]
            if range_title == '07单位': 
                list(temp_sr.replace(to_replace = '[单|位|数|量]',value='',regex=True).values[0])   #把所获得的字符串拆分 如"个个"拆为单个的字
            else:
                result_list = temp_sr.to_list()
            result_sr = pd.Series(data = result_list, name = range_title)
            if range_title == '01票号':
                data_center_h = (temp_df['luh'].values[0] + temp_df['ldh'].values[0]) //2
                get_h_adjust = int(crop_center_h - data_center_h)  #计算微调的高度系数,只能是整型
    if reg_type == 'contains':
        content_result = df['content'].str.contains(reg)
        if 'center_limit' in loc_dict: #如果字典有中心位置限制条件数据，则核对数据中心点位置是否符合条件
            center_df = df #赋值给中间临时表center_df，用以计算中心位置是否满足限制条件
            center_df[['luw','ruw','luh','ldh']].astype(int)
            center_df['center_w'] = (center_df['luw'] + center_df['ruw']) //2 
            center_df['center_h'] = (center_df['luh'] + center_df['ldh']) //2
            [center_w_min, center_w_max, center_h_min, center_h_max] = loc_dict['center_limit']
            cond_center = (center_w_min <= center_df.loc[:,'center_w']) & (center_df.loc[:,'center_w'] <= center_w_max) & \
                (center_h_min <= center_df.loc[:,'center_h']) & (center_df.loc[:,'center_h'] <= center_h_max) #坐标限定 
            content_result = content_result & cond_center
        if range_title == '07单位':
            cond_special = ~df['content'].str.contains('单\s*位|数\s*量') #不包含‘单位’字样
            content_result = content_result & cond_special
        content_df = df.loc[content_result,:]
        if range_title == '01票号':
            data_center_h = (content_df['luh'].values[0] + content_df['ldh'].values[0]) //2
            get_h_adjust = int(crop_center_h - data_center_h)  #计算微调的高度系数,只能是整型
        temp_df = content_df.loc[:,['content']]
        if not temp_df.empty:
            temp_sr = temp_df.iloc[:,0]
            if range_title == '07单位': 
                result_list = list(temp_sr.replace(to_replace = '[单|位|数|量]',value='',regex=True).values[0])  #把所获得的字符串拆分 如"个个"拆为单个的字
            else:
                result_list = temp_sr.to_list()
            result_sr = pd.Series(data = result_list, name = range_title)

    result_series_orderdic[range_title] = result_sr
    return result_series_orderdic, get_h_adjust


# ------------------------------func:Loc_jpg_content-----------------------
# 功能：根据条件在发票图片预识别表中直接查找文本
def Loc_jpg_content(df, cond_list, order_dict):
    range_title = cond_list[0]
    loc_method = cond_list[1]
    reg_type = cond_list[2]      
    reg = cond_list[3]
    count_limit = cond_list[4]
    loc_dict = cond_list[5]
    w_min, w_max, h_min, h_max = loc_dict['direct'][0], loc_dict['direct'][1], loc_dict['direct'][2], loc_dict['direct'][3]
    # # 调试代码：
    # if range_title == '10税前' and '0005.jpg' in file_path: 
    #      print(range_title,'\n', "w_limit:", w_min,w_max,'\n', 'h_limit:', h_min,h_max,'\n')
    result_sr = pd.Series(name = range_title) #结果初始化
    loc_tuple = [] #结果初始化
    if reg_type == 'extract':
        temp_df = df['content'].str.extract(reg)
        if len(temp_df) > 0:
            temp_df[['center_w','center_h','luw','ruw','luh','ldh']] = df[['center_w','center_h','luw','ruw','luh','ldh']]
            content_result = (temp_df.iloc[:,0].str.len() > 0)
            cond_loc=(w_min <= temp_df.loc[:,'center_w']) & (temp_df.loc[:,'center_w'] <= w_max) & \
                (h_min <= temp_df.loc[:,'center_h']) & (temp_df.loc[:,'center_h'] <= h_max) #坐标限定 
            cond_result = content_result & cond_loc #限定合并
            temp_cond_pandas = temp_df.loc[cond_result,:]
            if not temp_cond_pandas.empty:
                result_sr = temp_cond_pandas.iloc[:,0] #先赋值给result_sr，如果下面备用位置未取到，则就取当前赋的值
                loc_tuple = temp_cond_pandas.loc[:,['luw','luh']].values[0]
            if len(result_sr) == 0:
                #使用备用位置
                if len(loc_dict['direct'])>=8:
                    w_min, w_max, h_min, h_max = loc_dict['direct'][4], loc_dict['direct'][5], loc_dict['direct'][6], loc_dict['direct'][7]
                    #使用备用位置坐标识别
                    temp_df = df['content'].str.extract(reg)
                    temp_df[['center_w','center_h']] = df[['center_w','center_h']]
                    content_result = (temp_df.iloc[:,0].str.len() > 0)
                    cond_loc=(w_min <= temp_df.loc[:,'center_w']) & (temp_df.loc[:,'center_w'] <= w_max) & \
                        (h_min <= temp_df.loc[:,'center_h']) & (temp_df.loc[:,'center_h'] <= h_max) #坐标限定 
                    cond_result = content_result & cond_loc#限定合并
                    temp_cond_pandas = temp_df.loc[cond_result,:]
                    result_sr = temp_cond_pandas.iloc[:,0]
                    loc_tuple = temp_cond_pandas.loc[:,['luw','luh']].values[0]

            elif len(result_sr) >=1 and count_limit == '1':
                temp_cond_pandas = temp_df.loc[cond_result,:]
                result_sr = temp_cond_pandas.iloc[:,0].head(1)
                loc_tuple = temp_cond_pandas.loc[:,['luw','luh']].values[0]
            else:               
                result_sr = temp_df.loc[cond_result,0]
                loc_tuple = temp_cond_pandas.loc[:,['luw','luh']].values[0]

    elif reg_type == 'contains':
        content_result = df['content'].str.contains(reg)
        temp_df = df.loc[content_result,:]
        if len(temp_df) > 0:  #如果有结果
            cond_loc = (w_min <= temp_df.loc[:,'center_w']) & (temp_df.loc[:,'center_w'] <= w_max) & \
                (h_min <= temp_df.loc[:,'center_h']) & (temp_df.loc[:,'center_h'] <= h_max)  #坐标限定
            cond_result = content_result & cond_loc
            temp_cond_pandas = temp_df.loc[cond_result,:]
            if not temp_cond_pandas.empty:
                result_sr = temp_cond_pandas.iloc[:,0].head(1)
                loc_tuple = temp_cond_pandas.loc[:,['luw','luh']].values[0]
            else: 
                #使用备用位置
                if len(loc_dict['direct'])>=8:
                    w_min, w_max, h_min, h_max = loc_dict['direct'][4], loc_dict['direct'][5], loc_dict['direct'][6], loc_dict['direct'][7]
                    #使用备用位置坐标识别
                    content_result = df['content'].str.contains(reg)
                    temp_df = df.loc[content_result,:]

                    cond_loc = (w_min <= temp_df.loc[:,'center_w']) & (temp_df.loc[:,'center_w'] <= w_max) & \
                        (h_min <= temp_df.loc[:,'center_h']) & (temp_df.loc[:,'center_h'] <= h_max)  #坐标限定
                    cond_result = content_result & cond_loc
                    temp_cond_pandas = temp_df.loc[cond_result,:]
                    result_sr = temp_cond_pandas.iloc[:,0]
                    loc_tuple = temp_cond_pandas.loc[:,['luw','luh']].values[0]
                #记录再次精确识别的坐标：汉字金额左上角的w和h，传递给主程序  大写金额|{'trad'：(x,x,x)} 

    #---------------可在此处调试：上行加断点--------------------------------
    result_list = result_sr.to_list() #当前结果转为列表
    #---------------可在此处调试：上行加断点--------------------------------
    order_dict[range_title] = [pd.Series(result_list, name=range_title), loc_tuple] #追加识别信息到字典
    
    return order_dict

#--------------------------func: Tele_inv_ocr()---------------------------------
def Tele_inv_ocr(ocr_engines, result_series_orderdic, inv_dict, file_path, excel_file_path, err_info, engine_switch = 0):
# 匹配电子发票信息

    df_org = pd.read_excel(excel_file_path, sheet_name=0,header=None,index_col=None,na_values='', keep_default_na=True, dtype=object) #读取表格
    df_org = df_org.fillna('')
    df_org = df_org.astype(str)
    '去多空格'
    df_org = df_org.replace(to_replace = '\\n|\s+',value=' ',regex=True)
    df_org = df_org.replace(to_replace = '^\s+',value='',regex=True)
    #字典替换字符串:
    # rep = {'\n':'',' ':''}
    # rep = dict((re.escape(k), v) for k, v in rep.items())
    # #print(rep)
    # #print(rep.keys())
    # pattern = re.compile("|".join(rep.keys()))
    # #print(pattern)
    # my_str = pattern.sub(lambda m: rep[re.escape(m.group(0))], words)

    df_new = pd.DataFrame(data='', index = df_org.index, columns=['content'])
    #合并df_org的每一列到新表
    for i in df_org.columns:df_new['content'] = df_new['content'] +'|'+ df_org[i]
    #去重复分隔号
    df_new = df_new.replace(to_replace = '\|+',value='|',regex=True)
    #去开头结尾分隔号
    df_new = df_new.replace(to_replace = '^\||\|+$',value='',regex=True)

    fp_mark = False    
    if len(df_new.loc[df_new['content'].str.contains('发票'),:]) >0:
        fp_mark = True
    
    # 判断结果
    if fp_mark == False:  #pdf发票无找到电子发票字样, 返回。
        err_info = 'inv character not found.'
        return result_series_orderdic, err_info, inv_dict

    known_dict = {}
    known_dict = Get_known_from_from_xls_image(excel_file_path, ocr_engines[engine_switch])
            
    #字段参数配置：
    range_list = [
        [
            #发票号码
            '01票号',  #0 区域代号
            ['direct'], #1 直接提取
            'extract', #2 提取方式
            ['发票号码[：|:]?\s*(\d+)'],   #3 正则表达式
            '1', #4 提取数量限制
        ], 
        [
            #发票代码
            '02代码',  #字段标题
            ['direct'], #1 前面先qrcode，通过字典的known获取，其次crop识别
            'extract', 
            ['发票代码[：|:]?\s*(\d+)'], 
            '1', #提取数量限制 
        ],
        [
            #开票日期
            '03日期',
            ['direct'], #数据提取方式列表：汉字金额和前面先qrcode的，通过字典的known获取，其次crop、direct
            'extract', 
            ['(\d{4}\s*年\s*\d{2}\s*月\s*\d{2}\s*日)'],
            '1', #提取数量限制
        ],
        [
            #买方名称
            '04购方',
            ['direct'],
            'extract',
            [
                '^购买方信息\|名称：(.+?) 统一社会信用代码/纳税人识别号：',
                '名\s*称：\s*(.+?)\s*纳税人识别号'
            ], #非贪婪匹配
            '1'
        ],
        [
            #买方税号
            '05购方税号',
            ['direct'],
            'extract', 
            [
                '购买[\D]+纳税人识别号：[\|\s]*([0-9A-Z]{18?})',#非贪婪
                '纳税人识别号：([a-zA-Z0-9]{18})',
                
            ], #非贪婪匹配
            '1'
        ],
        [
            #商品名称  ok
            '06品名',
            ['direct'],
            'extract', 
            [
                '^项目名称\s*(.+)合\s*计\|',
                '^项目名称\s*(.+)合|',
            ], #贪婪匹配
            '1'
        ],
        [
            #单位
            '07单位',
            ['direct'],
            'extract', 
            [
                '^([\u4e00-\u9fa5]+)[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+$',
                '\|单\s*([\u4e00-\u9fa5]+)\|位\|',
                '\|?单\s*\|?\s*价\s*\|?\s*([\u4e00-\u9fa5]{1,3})\s*[.0-9]+',
                '\|?单[\s\|]*价[\|\s]*([\u4e00-\u9fa5]{1,3})\s*[.0-9]+',
                '\|?单[\s\|]*位[\|\s]*([\u4e00-\u9fa5]{1,3})[\|\s]*数[\|\s]*量[\|\s]*[.0-9]+[\|\s]*单[\|\s]*价[\|\s]*[.0-9]+',
            ], #贪婪匹配
            'n'    
        ],
        [
            #数量
            '08数量',
            ['direct'],
            'extract', 
            [
                '^[\u4e00-\u9fa5]+[\|\s]*([.0-9]+)[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+$',
                '量\s*([.0-9]+)\s*\|单',
                '\|?单[\s\|]*价[\|\s]*[\u4e00-\u9fa5]{1,3}\s*([.0-9]+)',
                '量[\s\|]*单[\s\|]*价[\|\s]*([.0-9]+)\s+[.0-9]+',
                '([.0-9]+)[\s\|]+[.0-9]+[\s\|]+[.0-9]+[\s\|]+[.0-9]+[\s\|]+[.0-9]+'
            ], #贪婪
            'n'    
        ],
        [
            #单价
            '09单价',
            ['direct'],
            'extract', 
            [
                '^[\u4e00-\u9fa5]+[\|\s]*[.0-9]+[\|\s]*([.0-9]+)[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+$',
                '价\s*([.0-9]+)\s*\|金',
                '\|?单[\s\|]*价[\|\s]*[\u4e00-\u9fa5]{1,3}\s*[.0-9]+[\|\s]+([.0-9]+)',
                '量[\s\|]*单[\s\|]*价[\|\s]*[.0-9]+\s+([.0-9]+)',
                '[.0-9]+[\s\|]+([.0-9]+)[\s\|]+[.0-9]+[\s\|]+[.0-9]+[\s\|]+[.0-9]+'
            ], #贪婪
            'n'
        ],
        [
            #商品明细税前金额
            '10税前',
            ['direct'],
            'extract', 
            [
                '^[\u4e00-\u9fa5]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*([.0-9]+)[\|\s]*[.0-9]+[\|\s]*[.0-9]+$',
                '[率|\|]\s*([.0-9]+)\s+[0-9]{1,2}%[\||税]',
                '金\s*额\s*([.0-9]+)[\|\s]*税率\s*[.0-9]+%[\|\s]*税\s*额',
                '[.0-9]+[\s\|]+[.0-9]+[\s\|]+([.0-9]+)[\s\|]+[.0-9]+[\s\|]+[.0-9]+'

            ], #贪婪
            'n'
        ],
        [
            #税率
            '11税率',
            ['direct'],
            'extract', 
            [
                '^[\u4e00-\u9fa5]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*([.0-9]+)[\|\s]*[.0-9]+$',
                '[率|\|]\s*[.0-9]+\s+([0-9]{1,2}%)[\||税]',
                '金\s*额\s*[.0-9]+[\|\s]*税\s*率\s*([.0-9]+%)[\|\s]*税\s*额',
                '[.0-9]+[\s\|]+[.0-9]+[\s\|]+[.0-9]+[\s\|]+([.0-9]+)[\s\|]+[.0-9]+'
            ],  #非贪婪
            '1'
        ],
        [
            #商品明细税额
            '12税额',
            ['direct'],
            'extract', 
            [
                '^[\u4e00-\u9fa5]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*[.0-9]+[\|\s]*([.0-9]+)$',
                '税\s*[\|]?\s*额\s*[\|]?\s*([.0-9]+)',
                '[.0-9]+[\s\|]+[.0-9]+[\s\|]+[.0-9]+[\s\|]+[.0-9]+[\s\|]+([.0-9]+)'
            ],  #贪婪
            'n'    
        ],
        [
            #合计税前金额
            '13合计税前',
            ['direct'], #1 前面先qrcode，通过字典的known获取，其次识别
            'extract', 
            [
                '[¥￥](-?\d+\.\d{0,2})[\|\s][¥￥]',
                '^([.0-9]+)\|[.0-9]+$'
            ], #非贪婪
            '1'
        ],
        [
            #合计税额
            '14合计税额',
            ['direct'],
            'extract', 
            [
                '[¥￥]-?\d+\.\d+[\|\s][¥￥](-?\d+\.\d+)',
                '^[.0-9]+\|([.0-9]+)$'
            ], #非贪婪
            '1'
        ],
        [
            #合计总额小写
            '15总额',
            ['direct'],
            'extract', 
            [
                '（小写）[¥￥](.+)',
                '价税合计[\|\s]*[零壹贰叁肆伍陆柒捌玖拾佰仟亿角分圆整]{2,}[\|\s]*[¥￥]?([.0-9]+)$'
            ],
            '1'   
        ],   
        [
            #合计大写
            '16大写',
            ['direct'],
            'extract', 
            # '([\(\)（）\u4e00-\u9fa5]{8,30})', 
            [
                '^价税合计（大写）\|(.+)\|（小写）',
                '价税合计[\|\s]*([零壹贰叁肆伍陆柒捌玖拾佰仟亿角分圆整]{2,})'
            ],
            '1'
        ],
        [
            #销方名称
            '17销方',
            ['direct'],
            'extract', 
            # '([\(\)（）\u4e00-\u9fa5]{7,30})', 
            [
                '销售方信息\|名称：(.+?) 统一社会信用代码',
                '销售方\s*\|\s*名\s*称：\s*([\u4e00-\u9fa5]+)\s*纳税人识别号'
            ],  #非贪婪
            '1'
        ],
        [
            #销方税号
            '18销方税号',
            ['direct'],
            'extract',
            [
                '销售[\D]+纳税人识别号：[\|\s]*([0-9A-Z]{18})',
                '纳税人识别号：([a-zA-Z0-9]{18})'
            ],  #非贪婪
            '-1'
        ]
    ]

    result_series_orderdic = OrderedDict()

    for i, cond_list in enumerate(range_list):

        result_series_orderdic, err_info = Loc_tele_content(df_new, known_dict, cond_list, result_series_orderdic)
        
    
    return result_series_orderdic, err_info, inv_dict
    
# ------------------------------func:Loc_tele_content-----------------------
# 功能：根据条件在发票图片预识别表中直接查找文本
def Loc_tele_content(df, known_dict, cond_list, order_dict):
    range_title = cond_list[0]
    loc_method = cond_list[1]
    reg_type = cond_list[2]      
    reg = cond_list[3]
    count_limit = cond_list[4]
    
    known_sr = pd.Series(name = range_title) #结果初始化
    
    result_sr = pd.Series(name = range_title) #结果初始化

    err_info = ''
    # 从已知的known中取值
    if range_title in known_dict:
        #此处加断点调试
        #先设定为known_dict中的值
        known_sr = pd.Series(data= known_dict[range_title], name = range_title)

    if reg_type == 'extract':
        #此处加断点调试
        temp_cond_pandas = pd.DataFrame()
        # if range_title == '11税率':
        #    print(range_title)
        #此处注释为调试代码，遍历reg列表，尝试匹配，提取到就跳出匹配：
        for _, r in enumerate(reg):
            temp_df = df['content'].str.extract(r)
            cond_result = temp_df.iloc[:,0].str.len() > 0
            temp_cond_pandas = temp_df.loc[temp_df.iloc[:,0].str.len() > 0,:]
            if len(temp_cond_pandas)>0:
                break

        if len(temp_cond_pandas)>0:
            # result_sr = temp_cond_pandas.iloc[:,0] #先赋值给result_sr，如果下面备用位置未取到，则就取当前赋的值
            if count_limit == '1':
                result_sr = temp_cond_pandas.iloc[:,0].head(1)
            elif count_limit == '-1':
                if len(temp_cond_pandas) == 1:
                    result_sr = temp_cond_pandas.iloc[:,0].head(1)
                else:
                    # 如果多条数据，选倒数第一个
                    result_sr = temp_cond_pandas.iloc[:,0].tail(-1)                  
            else:
                result_sr = temp_df.loc[cond_result,0]
            #去掉首尾空格：
            result_sr = result_sr.replace(to_replace='^\s|\s$',value='',regex=True)
            
        if range_title =='13合计税前':
            if len(known_sr) > 0:
                if len(result_sr) == 0:
                    result_sr = known_sr
                else:
                    result_value = result_sr.values[0]
                    known_value = known_sr.values[0]
                    if result_value == known_value:
                        #针对有的发票二维码读出来的金额不是税前金额而是总额的情况。
                        #只有是税前金额的情况，才以known为准
                        result_sr = known_sr.copy()
        
        elif range_title =='15总额':
            if '13合计税前' in known_dict:
                #针对有的发票二维码读出来的金额不熟税前金额而是总额的情况,如前期上海良和的发票
                known_sr = pd.Series(data= known_dict['13合计税前'], name = range_title)
                if len(known_sr) > 0:
                    if len(result_sr) > 0:
                        if result_sr.values[0] == known_sr.values[0]:     
                                result_sr = known_sr.copy()

        elif range_title == '06品名':
            # 先处理数据中的空格（符合的留下，不符合的删掉）
            target_sr = result_sr.str.extractall('([\u4e00-\u9fa5]+\s+[\u4e00-\u9fa5]+)')
            if len(target_sr) > 0:
                #对提取的要替换sr重新赋索引
                target_sr.index = list(range(len(target_sr)))
                #构建要替换成的字符串新sr
                replace_sr = target_sr.replace('\s+','',regex=True)
                #sr替换字符串
                new_sr = result_sr.copy()
                for i in enumerate(target_sr.index):
                    new_sr=new_sr.replace(target_sr.iloc[i],replace_sr.iloc[i],regex=True)
                result_sr = new_sr.copy()
            # 多品名按空格分割为多行:
            data = result_sr.iloc[0]
            if data.count(' ')>0:
                result_sr = pd.Series(data = data.split(' '),name=range_title)
        else:
            # 对于其他字段，如果result_sr没有值而known_sr有值，则以known_sr为准
            if len(result_sr) == 0 and len(known_sr) > 0:
                result_sr = known_sr.copy()

    #---------------可在此处调试：上行加断点--------------------------------
    # result_list = result_sr.to_list() #当前结果转为列表
    result_sr.name = range_title
    result_sr.index = list(range(len(result_sr))) #重新赋索引
    #---------------可在此处调试：上行加断点--------------------------------
    # order_dict[range_title] = [pd.Series(result_list, name=range_title)] #追加识别信息到字典
    order_dict[range_title] = [result_sr]

    return order_dict, err_info

def Get_known_from_from_xls_image(origin_pdf_xls_path, paddle_ocr):
    # 功能： 从pdf转换的xls中，识别其中包含的图片，生成known_dict
    xls_file_path = origin_pdf_xls_path
    # 解压目录
    pth_split = os.path.split(xls_file_path)
    pr = pth_split[0]
    nm = pth_split[1]

    nm_split = os.path.splitext(nm)
    fr = nm_split[0]
    ex = nm_split[1]

    unzip_path = os.path.join(pr, fr)
    
    sub_img_path = os.path.join(unzip_path, "xl\\media")

    result_title=['content']
    result_df = pd.DataFrame(columns = result_title)

    known_dict = {}

    draw_result_out = True

    wb = load_workbook(xls_file_path)
    ws = wb['Table 1']

    # for image in ws._images:
    #     # 输出图片的位置信息
    #     print("image:",image.path)
    #     print("anchor:",image.anchor._from)
    #     print((image.anchor._from.row,image.anchor._from.col),(image.anchor._from.rowOff,image.anchor._from.colOff), '\n')
        
    if not os.path.exists(unzip_path):
        os.mkdir(unzip_path)
    
    if draw_result_out == True:
        draw_result_folder = os.path.join(unzip_path, 'draw_result')
        if not os.path.exists(draw_result_folder):
            os.mkdir(draw_result_folder)

    with ZipFile(xls_file_path) as f:
        for file in f.namelist():
            # 解压图片部分的文件
            tempimg_path = ''
            if file.startswith("xl/media"):
                f.extract(file, path=unzip_path)    
                # 此处tempimg_path表达式和下面的os.path.join(sub_img_path, filename) 结果是相同的
                temp_img_name = os.path.split(file)[1]
                temp_img_fr = os.path.splitext(temp_img_name)[0]
                ext = os.path.splitext(temp_img_name)[1].lower()
                tempimg_path = os.path.join(unzip_path, file)

    #         if ext not in ['.jpg','.png','jpeg']:
    #             continue
    #         tempimg_path = os.path.join(sub_img_path, filename)        
            
    #         if os.path.exists(tempimg_path):

                #直接cv2方式:
                img = cv_imread(tempimg_path)

                #先判断图片是否为二维码，根据图片长款是否一致且大于300
                # print(img.shape)
                (h, w, _) = img.shape
                
                if 80 <= max(h, w) <= 200 and h == w:
                    #如果图片高宽一致且大于300，可能为二维码，尝试读取
                    codedata = pyzbar.decode(img)
                    if len(codedata) > 0:
                        data_str = codedata[0].data.decode()
                        if len(data_str) > 20:
                            data_list = data_str.split(',')
                            if len(data_list) > 4:
                                known_dict['01票号'] = data_list[3],
                                known_dict['02代码'] = data_list[2],
                                known_dict['03日期'] = data_list[5],
                                known_dict['13合计税前'] = data_list[4]

                # img_linear = cv2.resize(img, (img.shape[1]*4, img.shape[0]*4), cv2.INTER_LINEAR)
                # img_nearest = cv2.resize(img, (img.shape[1]*4, img.shape[0]*4), cv2.INTER_NEAREST)
                #              
                # 只识别高度像素在50以内的图片：
                if h < 50: 
                    enlarge = 4 #放大4倍

                    img_new = new(img, enlarge)

                    edge = 20
                    color = (255,255,255) #白色
                    img_large = cv2.copyMakeBorder(img_new,edge,edge,edge,edge, cv2.BORDER_CONSTANT,value=color) 
                    
                    enlarge_img_folder = os.path.join(unzip_path, 'img_enlarge')
                    if not os.path.exists(enlarge_img_folder):
                        os.mkdir(enlarge_img_folder)
                    enlarge_img_path = os.path.join(enlarge_img_folder, 'enlarge_' + temp_img_name)
                    cv2.imencode(".jpg", img_large)[1].tofile(enlarge_img_path)

                    result = paddle_ocr.ocr(img_large, cls=True)  #识别图像----------------
                    if len(result) > 0:
                        df = pd.DataFrame(data=[result[i][1][0] for i in range(len(result))],columns = result_title)
                        result_df = Collect_df(result_df, df)

                        if draw_result_out == True:
                            # draw result
                            from PIL import Image

                            image = Image.open(enlarge_img_path).convert('RGB')
                            # image = cv2_pil(img_large)
                            boxes = [line[0] for line in result]
                            txts = [line[1][0] for line in result]
                            scores = [line[1][1] for line in result]
                            im_show = draw_ocr(image, boxes, txts, scores, font_path='./fonts/simfang.ttf')
                            im_show = Image.fromarray(im_show)
                            # if range_title =='':
                            draw_result_name = 'draw_' + temp_img_name
                            # else:
                                # draw_result_name = 'draw_result_' + fr + '_' + range_title + ex 
                            draw_result_path = os.path.join(draw_result_folder, draw_result_name)
                            im_show.save(draw_result_path)

        temp_df = result_df.loc[:,'content'].str.extract('[¥￥]([.0-9]+)')
        temp_df.columns=['content']
        amount_df = temp_df.loc[temp_df['content'].notna(),:]
        if len(amount_df) >= 3:
            sqhj = float(known_dict['13合计税前'])
            amount_df = amount_df.astype(float)
            if sqhj > 1:
                values = amount_df.loc[amount_df['content']!=sqhj,'content'].values
                known_dict['15总额'] = max(values)
                known_dict['14合计税额'] = min(values)

        temp_df = result_df.loc[:,'content'].str.extract('^(91\S{16})$')
        temp_df.columns=['content']
        tax_numbers_df = temp_df.loc[temp_df['content'].notna(),:]
        if len(tax_numbers_df) > 0:
            our_number = ''
            known_dict['05购方税号'] = our_number
            values = tax_numbers_df.loc[tax_numbers_df['content']!=our_number,'content'].values
            if len(values)>0:
                known_dict['18销方税号'] = values[0]

        # print(result_df)
        # print('known_dict:',known_dict)

        img_ocr_result_folder = os.path.join(unzip_path, 'result')
        if not os.path.exists(img_ocr_result_folder):
            os.mkdir(img_ocr_result_folder)

        img_ocr_result_name = temp_img_fr + '.xlsx'
        img_ocr_result_path = os.path.join(img_ocr_result_folder, img_ocr_result_name)
        result_df.to_excel(img_ocr_result_path)
    
    return known_dict

#-----------------------------func: Pdf_tans_to()------------------------------------
def Pdf_tans_to(file_path, pdf_trans_to_file_path, trans_type = '.xlsx', temp_pdf_trans_excel_out = True) :
    # 可提取文字的pdf文件转为excel：
    # 先引入winerror、win32模块
    import winerror
    from win32com.client.dynamic import ERRORS_BAD_CONTEXT, Dispatch

    ERRORS_BAD_CONTEXT.append(winerror.E_NOTIMPL)
    output_folder_path = os.path.split(pdf_trans_to_file_path)[0]
    if not os.path.exists(output_folder_path):
        Create_clear_dir(output_folder_path)

    if trans_type == '.xlsx':
        trans_engion = 'com.adobe.acrobat.xlsx'
    elif trans_type == '.txt':
        trans_engion = 'com.adobe.acrobat.plain-text'
    else:
        trans_engion = 'com.adobe.acrobat.plain-text'
        
    try:
        AvDoc = Dispatch("AcroExch.AVDoc")    

        if AvDoc.Open(file_path, ""):            
            pdDoc = AvDoc.GetPDDoc()
            jsObject = pdDoc.GetJSObject()
            jsObject.SaveAs(pdf_trans_to_file_path, trans_engion)

    except Exception as e:
        print(str(e))

    finally:        
        AvDoc.Close(True)
        jsObject = None
        pdDoc = None
        AvDoc = None

    if os.path.exists(pdf_trans_to_file_path):
        return pdf_trans_to_file_path
    else:
        return None

#-----------------------------func: Pdf_tans_jpg()------------------------------------
def Pdf_tans_jpg(file_path, pdf_trans_jpg_file_path, temp_pdf_trans_jpg_out = True) :
    # 可提取文字的pdf文件转为excel：
    # 先引入winerror、win32模块

    output_folder_path = os.path.split(pdf_trans_jpg_file_path)[0]
    if not os.path.exists(output_folder_path):
        Create_clear_dir(output_folder_path)

    doc = fitz.open(file_path)
    pdf_name = os.path.splitext(file_path)[0]
    for pg in range(doc.pageCount):
        page = doc[pg]
        rotate = int(0)
        # 每个尺寸的缩放系数为2，这将为我们生成分辨率提高四倍的图像。
        zoom_x = 2.0
        zoom_y = 2.0
        trans = fitz.Matrix(zoom_x, zoom_y).preRotate(rotate)
        pm = page.getPixmap(matrix=trans, alpha=False)
        pm.writePNG(pdf_trans_jpg_file_path)

    if os.path.exists(pdf_trans_jpg_file_path):
        return pdf_trans_jpg_file_path
    else:
        return None

#------------------------------func: Check_result()----------------------------------------
def Check_result(result_pandas): #->Dataframe
    if len(result_pandas) == 0:
        return result_pandas
    # 整表转换为字符格式:
    edit_pandas = result_pandas.copy()
    edit_pandas = edit_pandas.fillna('')
    edit_pandas = edit_pandas.astype(str)
    temp_title_list = edit_pandas.columns.tolist()
    edit_pandas['err_info'] = ''  #清空err_info列
    pandas_title_list = edit_pandas.columns.tolist()
    inv_title_list = pandas_title_list[0:-2]
    detail_title_list = ['06品名','07单位' ,'08数量','09单价','10税前','12税额']
    num_title_list = ['08数量','09单价','10税前','11税率','12税额',\
        '13合计税前','14合计税额','15总额']
    one_row_title_list = ['01票号','02代码','03日期','04购方','05购方税号','13合计税前','14合计税额','15总额','16大写','17销方','18销方税号']
    one_row_title_list.sort()  #注意集合打乱了顺序，需要对标题列表重新排序
    #正则替换数字型文本字段区域的货币字符、百分号、括号和中文字符:
    edit_pandas.loc[:,num_title_list] = \
        edit_pandas.loc[:,num_title_list].replace(to_replace = '[￥¥%\s（）\(\)\u4e00-\u9fa5]',value='',regex=True)
    edit_pandas.loc[:,num_title_list] = \
        edit_pandas.loc[:,num_title_list].replace(to_replace = '[:：]',value='.',regex=True)
    edit_pandas.loc[:,'05购方税号'] = \
        edit_pandas.loc[:,'05购方税号'].replace(to_replace = '[:：]',value='',regex=True)
    # 替换 品名标点字符：
    edit_pandas.loc[:,'06品名'] = \
        edit_pandas.loc[:,'06品名'].replace(to_replace = '^[米水冰]|[\+\*#]',value=' ',regex=True)
    edit_pandas.loc[:,'06品名'] = \
        edit_pandas.loc[:,'06品名'].replace(to_replace = '^\s',value='',regex=True)
    # 字段修正：公司名错别字：
    comp_dict = {
        '有限公司' : '有限公司',
        }
    edit_pandas = edit_pandas.replace({'17销方':comp_dict})
    # 字段修正 通过文件读入字典修正
    replace_file = 'D:\\pyscripts\\发票修正.xlsx'
    if os.path.exists(replace_file):
        replace_df = pd.read_excel(replace_file, sheet_name=0,header=0, keep_default_na=True, dtype=object) #读取表格
        if not replace_df.empty:
            replace_df = replace_df.fillna('')
            edit_df_title_list = edit_pandas.columns.to_list()
            replace_df_title_list = replace_df.columns.to_list()
            for _, title in enumerate(replace_df_title_list):
                if title in edit_df_title_list:
                    if not replace_df.loc[replace_df[title]!='',:].empty:
                        #如果replace_df里对应edit_df的相应字段不为空
                        replace_title = title + '修正'
                        if replace_title in replace_df_title_list:
                            #如果有相应字段的修正列,则遍历字段列，用修正列替换
                            for _, row in enumerate(replace_df[[title,replace_title]].iterrows()):
                                str_origin = row[1].values[0]
                                str_replace = row[1].values[1]
                                edit_pandas[title] = edit_pandas[title].replace(to_replace = str_origin, value=str_replace, regex=True)
    # 获得遍历需要的发票起止行
    row_start_index = edit_pandas.loc[edit_pandas['file_path'].str.len()>0,'file_path'].index
    row_start_list = row_start_index.to_list()
    temp_index = row_start_index - 1
    temp_list = temp_index.to_list()
    row_end_list = temp_list[1:]
    row_pandas_last = edit_pandas.index[-1]
    row_end_list.append(row_pandas_last)
    rows_tuple = zip(row_start_list,row_end_list)

    for i, (row_start, row_end) in enumerate(rows_tuple):
        err_info = ''
        err_blank = ''
        err_code = ''
        err_product = ''
        err_num = ''
        this_inv_pandas = edit_pandas.iloc[row_start:row_end+1, :] #截取单张发票的数据行区域到one_inv_pandas   
        # file_path = this_inv_pandas.loc[this_inv_pandas.index[0], 'file_path'] #此处调试：
        # if '\\23.jpg' in file_path:
        #     print(file_path)
        #数值核对
        num_extract_reg = '((?:\d+|\d{0,3}(?:,\d{3})*)\.?\d{0,})\s*$'
        # 如果数字列包含除了点以外的非数字字符，则去除非数字字符
        for _, num_title in enumerate(num_title_list):
            this_inv_pandas.loc[:,num_title] = this_inv_pandas.loc[:,num_title].str.extract(num_extract_reg)
            this_inv_pandas.loc[:,num_title_list]=this_inv_pandas.loc[:,num_title_list].replace('^$','0',regex=True)
        this_inv_pandas.loc[:,num_title_list] = this_inv_pandas.loc[:,num_title_list].astype(float)   
        #1. 税率换算小数
        if this_inv_pandas.loc[:,'11税率'].values[0] >1:
            this_inv_pandas.loc[:,'11税率'] = this_inv_pandas.loc[:,'11税率']/100
        # 税前合计
        num_sum_pretax_amount = round(sum(this_inv_pandas['10税前'].values),2)
        num_total_pretax_amount = this_inv_pandas['13合计税前'].values[0]
        # 累计税额和税额合计
        num_total_tax = this_inv_pandas['14合计税额'].values[0]
        num_sum_detail_tax = round(sum(this_inv_pandas['12税额'].values), 2)
        # 税前合计+税额合计 和 发票金额 
        num_total_amount= this_inv_pandas['15总额'].values[0]
        sum_total = num_total_pretax_amount + num_total_tax

        #检查空白区域：
        title_blank_list = []
        err_inv_list = []
        
        for _, title in enumerate(detail_title_list):
            cond1 = this_inv_pandas.loc[:, title] == ''
            cond2 = this_inv_pandas.loc[:, title] == 0
            cond = cond1 | cond2
            count_blank = len(this_inv_pandas.loc[cond,:])
            if count_blank > 0:
                #如果有空值
                title_blank_list.append(title)
            if title == '06品名':
                cond = this_inv_pandas.loc[:, title].str.contains('品[\u4e00-\u9fa5]')
                product_wrong_df = this_inv_pandas.loc[cond,'06品名']
                count_product_err = len(product_wrong_df)
                if count_product_err > 0:
                    err_product = err_product + 'Check product name:' + ','.join(product_wrong_df.to_list()) + '.'
                if '品名' not in err_blank:
                    if len(this_inv_pandas.loc[~this_inv_pandas['06品名'].str.contains('[\u4e00-\u9fa5]\s[\u4e00-\u9fa5]'),:]) > 0:
                        # 如果品名不符合"若干汉字+空格+若干汉字"格式，提示错误
                        err_product = err_product + '品名格式不符“类品+空格+品名”.'
        for _, title in enumerate(one_row_title_list):
            if title == '发票号码':
                temp_df = this_inv_pandas.loc[this_inv_pandas['file_path']!='', '发票号码']
                temp_df['发票号长度'] = temp_df['发票号'].apply(lambda x:len(x))
                temp_check_df = temp_df.loc[~((temp_df['发票号长度']==8) |(temp_df['发票号长度']==20)),: ]
                if len(temp_check_df) > 0:
                    err_inv_list.append('Inv number lenth illegal')
                temp_check_df= temp_df.loc[temp_df['发票号'].str.contains('\D'), :]
                if len(temp_df) > 0:
                    err_inv_list.append('Inv number character illegal')
                
            cond1 = this_inv_pandas.loc[this_inv_pandas.index[0], title] == ''
            cond2 = this_inv_pandas.loc[this_inv_pandas.index[0], title] == 0
            cond = cond1 | cond2
            if cond == True:  #即为空
                
                if title == '02代码':
                    if len(this_inv_pandas.loc[this_inv_pandas.index[0], '01票号']) == 20: #表示电子发票，可以没有代码
                        continue
                if title == '15总额':
                    #如果是总额字段，是0，如果大写不为空，用大写转换为小写，替换总额
                    txt = this_inv_pandas.loc[this_inv_pandas.index[0], '16大写']
                    if not txt == '':
                        trad = txt.split('|')[0]
                        repl_dict = {
                            '参' : '叁',
                            '柴' : '柒',
                            '什' : '仟'
                            }
                        trad = repl_by_dict(trad, repl_dict)
                        money = trad_to_int(trad)
                        if not money == trad:
                            money = float(money)
                            if money > 0:
                                this_inv_pandas.loc[this_inv_pandas.index[0], title] = money
                                continue
                    else: #如果大写为空的话，再看合计税前和合计税额，如果都有数值，则用两个金额的和替换发票总额
                        if num_total_pretax_amount >0 and num_total_tax > 0:
                            this_inv_pandas.loc[this_inv_pandas.index[0], title] = sum_total
                            continue
                if title == '16大写':
                    continue
                
                title_blank_list.append(title)

         #如果公司税号不存在空值，进行校验：
        if '05购方税号' not in title_blank_list:
            if this_inv_pandas['05购方税号'].values[0] != user_code:
                err_code = '购方税号['+ this_inv_pandas['05购方税号'].values[0] + ']不是“' + user_code + '”。'

        if len(title_blank_list) > 0:
            title_blank_list.sort()
            err_blank = 'Null:[' + ','.join(title_blank_list) + ']。' #标识到错误记录
                    
        diff_pretax_amount = round(num_total_pretax_amount - num_sum_pretax_amount, 2)
        if diff_pretax_amount != 0:
            err_num = err_num + '税前之和≠合计税前[' + str(num_total_pretax_amount) + ' - ' + str(num_sum_pretax_amount) + ' = ' + str(diff_pretax_amount) + ']。'
        
        # 2 税前合计+税额合计 和 发票金额 , 累计税额和税额合计
        sum_total_pretax_tax = round(num_total_pretax_amount + num_total_tax, 2)
        diff_total = round(num_total_amount - sum_total_pretax_tax, 2)
        diff_tax = round(num_total_tax - num_sum_detail_tax, 2)
        if diff_total != 0:
            err_num = err_num + '税前合计与税额合计之和≠发票总额[' + str(sum_total_pretax_tax) + '≠' + str(num_total_amount) + ']。'
        if diff_tax != 0:
            err_num = err_num + '明细税额之和≠14合计税额:[' + str(num_sum_detail_tax) + ' ≠ ' + str(num_total_tax) +']。'

        # 3 数量*单价=税前金额
        quantity_price_df = this_inv_pandas.loc[:,['06品名','08数量','09单价','10税前']]
        quantity_price_df['diff_quantity_price'] = quantity_price_df['08数量'] * quantity_price_df['09单价'] - quantity_price_df['10税前']
        round_quantity_price_df = quantity_price_df.loc[:,'diff_quantity_price'].astype(float).round(2)
        quantity_price_df['diff_quantity_price'] = round_quantity_price_df
        diff_quantity_price_df = quantity_price_df.loc[quantity_price_df['diff_quantity_price'] != 0,:]
        if not diff_quantity_price_df.empty:
            str_temp_quantity      = '，'.join(diff_quantity_price_df['08数量'].astype(str).tolist())
            str_temp_price         = '，'.join(diff_quantity_price_df['09单价'].astype(str).tolist())
            str_temp_pretax_amount = '，'.join(diff_quantity_price_df['10税前'].astype(str).tolist())
            str_temp_diff          = '，'.join(diff_quantity_price_df['diff_quantity_price'].astype(str).tolist())

            err_num = err_num + '量*价≠税前,差异明细：['+ str_temp_quantity + ']×['+ str_temp_price + ']-['+  str_temp_pretax_amount + ']=[' + str_temp_diff + ']。'
        err_inv = '票号格式错误['+','.join(err_inv_list) + ']。' if len(err_inv_list)>0 else ''
        err_info = err_inv + err_blank + err_code + err_product + err_num
        err_before = this_inv_pandas.loc[:,'err_info'].values[0]
        err_info = err_before + err_info
        this_inv_pandas.loc[this_inv_pandas.index[0],'err_info'] = err_info
        edit_pandas.iloc[row_start:row_end + 1,:] = this_inv_pandas #temp_pandas的写回结果到截取区域
    
    result_pandas = edit_pandas # 修改数据edit_pandas写回result_pandas
  
    return result_pandas


#---------------输出结果----------------
def Log_result_file(result_pandas,result_file_path,result_sheet_name):
    writer = pd.ExcelWriter(result_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    result_pandas.to_excel(writer,sheet_name=result_sheet_name,index=False)
    writer.close()
    return True

#---------------添加超链接---------------
def Add_hyperlink(result_file_path,result_sheet_name):
#添加文件路径超链接
    wb = load_workbook(result_file_path)
    wb.move_sheet(result_sheet_name, offset=-1)
    ws = wb[result_sheet_name]
    wb._active_sheet_index = 0  #激活第一个工作表
    rng = ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=19, max_col=20)
    for col in rng:	#列方法
        for cell in col:	#遍历列
            txt = cell.value
            if txt is None:
                continue
            if len(txt) > 0:
                if cell.column == 19:
                    pr,nm,fr,ex = pathsplit(txt)
                    # 绝对路径
                    # cell.hyperlink = 'file:\\' + txt
                    # 相对路径
                    cell.hyperlink = '..\\' + nm
                    cell.font = Font(color=colors.Color(index = 4), italic=True)
                else:
                    cell.font = Font(color=colors.Color(index = 2), italic=False)
    wb.save(result_file_path)
    ws = None
    wb = None
    return True

def Collect_df(collect_df, item_df):
    #汇总df表
    if len(item_df) == 0:
        return collect_df
    
    if collect_df.empty:
        collect_df = item_df
    else:
        test_set = {0,1}
        collect_df_col_set = set(collect_df.columns)
        item_df_col_set = set(item_df.columns)
        if len(collect_df_col_set - item_df_col_set) > 0:
            # 如果合并表和被合并子表列名不一致，则忽略索引后合并
            temp_collect_df = collect_df.copy()
            temp_collect_title_list = temp_collect_df.columns.to_list()
            temp_collect_title_df = pd.DataFrame(data = temp_collect_title_list).T
            temp_collect_df.columns  = list(range(len(temp_collect_df.columns)))
            collect_df = pd.concat([temp_collect_title_df, temp_collect_df], ignore_index = True, axis = 0)
            
            temp_item_df = item_df.copy()
            temp_item_title_list = temp_item_df.columns.to_list()
            temp_item_title_df = pd.DataFrame(data = temp_item_title_list).T
            temp_item_df.columns  = list(range(len(temp_item_df.columns)))
            item_df = pd.concat([temp_item_title_df, temp_item_df], ignore_index = True, axis = 0)
            
            collect_col_num = len(temp_collect_title_list)
            item_df_col_num = len(temp_item_title_list)
            max_col = max(collect_col_num, item_df_col_num)
            collect_col_dif = max_col - collect_col_num
            item_col_dif = max_col - item_df_col_num
            if collect_col_dif > 0:
                for i in range(collect_col_num, max_col + 1):
                    temp_collect_df[i] = ''
            if item_col_dif > 0:
                for i in range(item_df_col_num, max_col + 1):
                    temp_item_df[i] = ''
            collect_df = temp_collect_df
            item_df = temp_item_df

        collect_df = pd.concat([collect_df, item_df], ignore_index = True, axis = 0)
        # 重设index从1 
        collect_df = reset_nature_index(collect_df)
    return collect_df
#----------------------------Log_df_to_file()----------------------
def Log_df_to_file(df, save_path, sheet_name, keep_exists = True):
    # 输出数据表：
    # 默认不删除原来文件表中内容，即默认累积keep_exists=True
    writer = pd.ExcelWriter(save_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    pandas_write = pd.DataFrame()
    if not df.empty:
        if keep_exists == True:
            # 如果累积原文件内数据
            df_title = df.columns.to_list()
            df_non_title = df
            df_non_title.columns  = list(range(len(df_non_title.columns)))            
            pandas_write = pd.read_excel(save_path, sheet_name=sheet_name,index_col=0,header = 0,keep_default_na=True,dtype=object) #读取表格
            pandas_write = Collect_df(pandas_write, df)
        else:
            # 如果不累积原文件内数据
            pandas_write = df
    if not pandas_write.empty:
        pandas_write.to_excel(writer,sheet_name=sheet_name)  
    writer.close()
    
    return True

# --------------------reset_int_index()----------------
def reset_nature_index(df): 
    # 重设自然数索引index从1开始
    df.index = list(range(1,len(df)+1))
    return df

# ------------------------------func:pil_enhance()-----------------------
def pil_enhance(img):
    # 功能图片预处理
    # 增加亮度
    img = ImageEnhance.Brightness(img).enhance(1.0)
    # 锐利化
    img = ImageEnhance.Sharpness(img).enhance(1.5)
    # 增加对比度
    img = ImageEnhance.Contrast(img).enhance(2.0)
    # 灰度化
    img_result = img.convert('L')

    return img_result

def new(img, enlarge):
    # 放大图像为enlarge倍
    img_new = np.zeros((img.shape[0] * enlarge, img.shape[1] * enlarge, img.shape[2]))
    for i in range(img.shape[0]):
        for j in range(img.shape[1]):
            for m in range(4):
                for n in range(4):
                    img_new[4*i + m][4*j + n] = img[i][j]
    return img_new

def Pil_make_border(image, edge = 20):
    # 图像扩充边界，扩充后边界增加 1/2 * edge 个像素
    iw, ih = image.size  # 原始图像的尺寸
    w, h = iw + edge, ih + edge  # 目标图像的尺寸
    target_size = (w, h)
    
    # scale = min(float(w) / float(iw), float(h) / float(ih))  # 转换的最小比例
    # 保证长或宽，至少一个符合目标图像的尺寸
    # nw = int(iw * scale)
    # nh = int(ih * scale)
    nw = iw
    nh = ih

    image = image.resize((nw, nh), Image.BICUBIC)  # 缩小图像
    # 白色
    color=(255,255,255) 

    new_image = Image.new('RGB', target_size, color)  # 生成白色色图像
    # // 为整数除法，计算图像的位置
    new_image.paste(image, ((w - nw) // 2, (h - nh) // 2))  # 将图像填充为中间图像，两侧为白色的样式
    # new_image.show()
    return new_image

#--------------------------func: pil_cv2()---------------------------   
# pil 转 cv2
def pil_cv2(pil_image):
    cv2_image = cv2.cvtColor(np.asarray(pil_image), cv2.COLOR_RGB2BGR)
    return cv2_image
#--------------------------func: cv2_pil()---------------------------   
# cv2 转 pil
def cv2_pil(img_cv):
    pil_image = Image.fromarray(cv2.cvtColor(img_cv,cv2.COLOR_BGR2RGB))
    return pil_image

#--------------------------func: cv_imread()---------------------------
# 程序功能：cv2读取图片
def cv_imread(file_path):
    cv_img = cv2.imdecode(np.fromfile(file_path,dtype=np.uint8),cv2.IMREAD_COLOR)
    return cv_img

#--------------------------func: trad_to_int()---------------------------
# 程序功能：大写金额转为小写数字
# 正则表达式法    
def trad_to_int(money):
    # 转换字典
    trad_dict = {'零':0,'壹':1,'贰':2,'叁':3,'肆':4,'伍':5,'陆':6,'柒':7,'捌':8,
    '玖':9,'拾':10,'佰':100,'仟':1000,'万':10000,'亿':100000000,'角':0.1,'分':0.01}

    trad = re.search(r"[零壹贰叁肆伍陆柒捌玖拾佰仟亿角分]+", money)
    if trad is not None:
        num = 0
        add = 0
        sum = 0
        for i in money:
            if i in ['零','壹','贰','叁','肆','伍','陆','柒','捌','玖']:
                add = trad_dict[i]
                sum = sum + add
            elif i in ['拾','佰','仟','亿','角','分']:
                num = add * trad_dict[i]
                sum = sum - add
                sum = sum + num
                add = num
            elif i == '万' or i == '亿':
                sum = sum * trad_dict[i]
        sum = str(sum)
        return sum
    else:
        return money
#-------------------------=func: Fill_na_result()-------------------------
def Fill_na_result(result_df):
    # 填充处理：务必先处理na值，再进行后续处理。
    result_df.loc[:,'03日期'] = result_df.loc[:,'03日期'].apply(lambda x: delta_date(x))
    result_df.loc[:,'11税率'] = result_df.loc[:,'11税率'].fillna(method='ffill')
    result_df.iloc[:,0:7] = result_df.iloc[:,0:7].fillna('')
    result_df.iloc[:,7:15] = result_df.iloc[:,7:15].fillna('0')
    result_df.iloc[:,15:] = result_df.iloc[:,15:].fillna('')
    result_df = result_df.fillna('')  

    return result_df

#---------------------------- delta_date():int值转date 文本格式-------------
def delta_date(para):
    time = para
    if isinstance(para,int):
        time = pd.to_datetime('1899-12-30') + pd.Timedelta(str(int(para))+'days')
        time = time.strftime("%Y-%m-%d")
    elif isinstance(para,float):
        time = ''
        # time = pd.to_datetime('1899-12-30') + pd.Timedelta(str(origin_pandas.iloc[0,2])+'days')
    return time

#--------------------------func: repl_by_dict()---------------------------
#字典替换字符串
def repl_by_dict(my_str,repl_dict):
    for (k,v) in repl_dict.items():
        my_str = my_str.replace(k, v)
    return my_str

#--------------------------func: pathsplit()---------------------------
# 路径分割
def pathsplit(f) ->tuple:
    parent = os.path.split(f)[0]
    fullname = os.path.split(f)[1]
    frontname = os.path.splitext(fullname)[0]
    extname = str.lower(os.path.splitext(f)[1])
    return (parent,fullname,frontname,extname)

#--------------------------func: Create_clear_dir()---------------------------
# 创建空目录
def Create_clear_dir(folder_path):
    if os.path.exists(folder_path):  #清空临时文件夹，如果不存在则新
        for dirpath, dirnames, filenames in os.walk(folder_path):
            for filepath in filenames: #清空临时txt文件夹
                delFolderorFile(os.path.join(dirpath, filepath))  
    else:
        os.mkdir(folder_path) #新建文件夹
    if os.path.exists(folder_path):
        return True
    else:
        return False
#----------------------------func: delFolderorFile()------------------------------
#删除目录或文件
def delFolderorFile(folder_path):
    if not os.path.exists(folder_path):
        return False
    if os.path.isfile(folder_path):
        os.remove(folder_path)
        return
    for m in os.listdir(folder_path):
        n = os.path.join(folder_path, m)
        if os.path.isdir(n):
            #递归调用delFolderorFile
            delFolderorFile(n)
        else:
            os.unlink(n)
    os.rmdir(folder_path) #删除空目录

#--------------------------func: cal_angle()--------------------------
#两点坐标计算角度弧度
def cal_angle(p1, p2):
    """ 
    px : (横坐标，纵坐标)
    """
    angle=math.atan2(p2[1]-p1[1], p2[0]-p1[0])
    # angle=math.degrees(angle)
    return angle * (180 / math.pi)

#--------------------------func: killexcel()--------------------------
def killexcel():
    pids = psutil.pids()
    for pid in pids:
        try:
            p = psutil.Process(pid)
            if str.upper(p.name()) == 'EXCEL.EXE':
                cmd = 'taskkill /F /IM EXCEL.EXE'
                os.system(cmd)
            # cmd中输入：tasklist /m exp*
        except Exception as e:
            print(e)
#----------------------------------------------------------------#
#                    --- MAIN PROGRAM ---                        #
#----------------------------------------------------------------#

if __name__ == '__main__':


    print('\n',datetime.now().strftime("%H:%M:%S"),'Program start running...\n')
    killexcel()

    user_name = ''
    user_code = ''

    reserve_template_before = True #默认为True：保留之前的中间文件,False:删除中间文件,全新运行。
    ocr_excel_out = True #True:输出临时excel文件
    draw_result_out = True #绘制识别结果
    enhance = False
    acumulate_input = 'y'  #默认增量运行
    #设置快慢速所用引擎: 0-快速,1-慢速,2-平衡
    prepare_engine = 1 
    precise_engine = 1 
    root = Tk()
    print('Please choose the images folder:')
    origin_folder_path = filedialog.askdirectory()
    if len(origin_folder_path) > 0:
        origin_folder_path = origin_folder_path.replace('/','\\')
        print(datetime.now().strftime("%H:%M:%S"),'The images folder you chose：', origin_folder_path)
    else:
        print(datetime.now().strftime("%H:%M:%S"),'No file chosen. \nQuit.')
        exit()
    root.destroy()

    result_folder_name = 'result' #结果文件夹
    result_sheet_name ='result' #结果工作表名
    result_folder_path = os.path.join(origin_folder_path, result_folder_name) #结果文件夹
    if not os.path.exists(result_folder_path):
        Create_clear_dir(result_folder_path) #建立干净的中间临时文件夹            
    
    result_file_name = 'result' + '.xlsx'
    result_file_path = os.path.join(result_folder_path, result_file_name) #结果文件路径

    run_renew = True

    pr,nm,fr,ex = pathsplit(result_file_path)
    now = datetime.now()
    back_str = now.strftime("%Y%m%d_%H%M%S")
    back_file_name = fr + '_' + back_str + ex
    back_file_path = os.path.join(result_folder_path, back_file_name) 
    origin_pandas = pd.DataFrame() #初始化
    t0 = datetime.now()
    if os.path.exists(result_file_path):

        print(datetime.now().strftime("%H:%M:%S"), f'Found previous result: {result_file_path} .')
        
        # 选择识别还是只是检查结果，默认识别。
        ocr_input = 'y'
        print('\nChoose please: \n"y" - run the orgnize engine.   "n" - only check the result, do not run engine.\n')
        ocr_input = input('Input(y/n):\n')
        
        # 如果只是检查结果，检查结果后退出:
        # 先初始化
        origin_pandas = pd.DataFrame()
        if str.lower(ocr_input) == 'n':
            #结果文件备份：
            shutil.copy(result_file_path, back_file_path)
            #获取结果文件
            try:
                origin_pandas = pd.read_excel(result_file_path, sheet_name=result_sheet_name,header=0, keep_default_na=True, dtype=object) #读取表格
            except ValueError:
                origin_pandas = pd.read_excel(result_file_path, sheet_name=0,header=0, keep_default_na=True, dtype=object) #读取表格
            #检查结果
            result_pandas = Check_result(origin_pandas)
            #添加超链接
            Log_result_file(result_pandas,result_file_path,result_sheet_name)
            Add_hyperlink(result_file_path,result_sheet_name)
            print('\n')
            print(datetime.now().strftime("%H:%M:%S"), 'Done.《', result_file_path, '》checked over.')
            # 退出程序
            exit()

        # 如果选择运行，指定增量运行还是全新运行 ,默认是增量运行
        if ocr_input.lower() == 'y':
            print('\nChoose run method: \n"y" - Run acumulated to the existed result.   \n"n" - Run fresh and delete all existed results and template folders.\n')
            acumulate_input = input('Input(y/n):\n')

        #如果增量运行，读取已存在结果
        if acumulate_input.lower() =='y':
            # 因是增量运行，所以把临时中间文件全新运行设置为否
            reserve_template_before = 'y'
            #结果文件备份：
            shutil.copy(result_file_path, back_file_path)
            #获取结果文件
            try:
                origin_pandas = pd.read_excel(result_file_path, sheet_name=result_sheet_name,header=0, keep_default_na=True, dtype=object) #读取表格
            except ValueError:
                origin_pandas = pd.read_excel(result_file_path, sheet_name=0,header=0, keep_default_na=True, dtype=object) #读取表格
            # df = pd.read_excel(os.path.join(init_dir, '地理区域.xlsx'), converters={'父级地理区域编号': str, '地理区域编号': str})
        else:
            #如果全新运行，删除结果文件
            Create_clear_dir(result_folder_path)

        if not origin_pandas.empty:
            origin_pandas = Fill_na_result(origin_pandas)
    
    temp_folder_name = 'temp' #中间临时文件夹
    temp_folder_path = os.path.join(origin_folder_path, temp_folder_name)
    
    if not(reserve_template_before) or not(os.path.exists(temp_folder_path)):
        # 如果不保留上次临时文件，或者临时文件夹不存在，则清空或建立新临时文件夹
        Create_clear_dir(temp_folder_path) #建立干净的中间临时文件夹     
    
    print(datetime.now().strftime("%H:%M:%S"),'Start the engine...')
    #识别引擎列表：
    mobile_ocr = hub.Module(name="chinese_ocr_db_crnn_mobile")  #0号引擎
    paddle_ocr = PaddleOCR(enable_mkldnn=True,use_angle_cls=True, lang='ch')       #1号引擎       
    # server_ocr = hub.Module(name="chinese_ocr_db_crnn_server")  #2号引擎
    
    ocr_engines = [mobile_ocr, paddle_ocr]  #引擎集合
    print(datetime.now().strftime("%H:%M:%S"),'Engine start running...')

    result_pandas_orderdic = OrderedDict() #定义pandas字典
    duplicate_sheet_name = 'duplicate'
    duplicate_pandas = pd.DataFrame() #初始化
    try:
        duplicate_pandas = pd.read_excel(back_file_path, sheet_name=duplicate_sheet_name,header=0, keep_default_na=True, dtype=object) #读取表格
    except:
        pass
    if duplicate_pandas.empty:
        duplicate_pandas.to_excel(result_file_path,sheet_name=duplicate_sheet_name,index=False) #如果没有重复发票表，则创建一个重复发票工作表
    cnt_file = len({p.resolve() for p in Path(origin_folder_path).glob("*") if p.suffix in [".jpg", ".pdf"]})
    cnt_done_pre = 0
    cnt_duplicate_pre = 0
    # 如果是增量运行，取已存在的识别文件数和重复文件数。否则都默认为0
    if acumulate_input.lower() =='y':
        if not origin_pandas.empty:
            cnt_done_pre = len(origin_pandas.loc[origin_pandas['file_path'].notnull(),:])
        if not duplicate_pandas.empty:
            cnt_duplicate_pre = len(duplicate_pandas.loc[duplicate_pandas['file_path'].notnull(),:])
    inv_dict = {} #发票号字典 {发票号：[文件名]}
    walk_folder_args = {'ocr_engines':ocr_engines, 'temp_folder_path': temp_folder_path, 'engine_switch':prepare_engine}
    #------------------识别核心函数--------------------------------
    result_pandas,duplicate_pandas = walk_folder_ocr(origin_pandas, duplicate_pandas, origin_folder_path,**walk_folder_args)
    #-------------------------------------------------------------
    print('\n')
    print(datetime.now().strftime("%H:%M:%S"),'Get the result.')
    cnt_done = 0
    cnt_duplicate = 0
    if not result_pandas.empty:
        cnt_done = len(result_pandas.loc[(result_pandas['file_path']!='') & (result_pandas['file_path'].notnull()),:]) - cnt_done_pre
    if not duplicate_pandas.empty:
        cnt_duplicate = len(duplicate_pandas.loc[(duplicate_pandas['file_path']!='') & (duplicate_pandas['file_path'].notnull()),:]) - cnt_duplicate_pre
    if not result_pandas.empty:
        print(datetime.now().strftime("%H:%M:%S"),'Checking result data...')
        result_pandas = Check_result(result_pandas)  #最后检查一遍结果,并重新写入结果文件
    Log_result_file(result_pandas,result_file_path,result_sheet_name)
    print(datetime.now().strftime("%H:%M:%S"),'Result data check over.')
    # 结果添加超链接
    Add_hyperlink(result_file_path,result_sheet_name)
    paddle_ocr = None
    server_ocr = None
    mobile_ocr = None
    ocr_engines = None
    print('\toriginal image path:  ' + origin_folder_path)
    print('\toutput file path:  ' + result_file_path)
    t1 = datetime.now()
    tx = t1 - t0 
    v = 0
    try:
        v = round(tx.total_seconds() / (cnt_done + cnt_duplicate) , 2)
    except:
        pass
    print('\n')
    print( t1,'\n Done. Time spent: ', str(tx).split('.')[0], '. Files total: ' + str(cnt_file) \
        + '. Already done before start: ' + str(cnt_done_pre) \
        + '. Already find duplicate before start: ' + str(cnt_duplicate_pre) \
        + '. \n Files recognized this time total: ' + str(cnt_done + cnt_duplicate) \
        + ', valid: ' + str(cnt_done) + ', duplicate: ' + str(cnt_duplicate) \
        + ', Everage: ' + str(v) + ' s.\n') 
    cnt_done_total = cnt_done_pre + cnt_done
    cnt_duplicate_total = cnt_duplicate_pre + cnt_duplicate
    if cnt_done_total != cnt_duplicate_total:
        print('Warning: 有效发票数:' + str(cnt_done_total) + ' 重复发票数：' + str(cnt_duplicate_total) + ', 检查是否有发票号码错误。')
